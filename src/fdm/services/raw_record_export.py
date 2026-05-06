from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
import math
import os
import posixpath
import re
import tempfile
import zipfile
from xml.etree import ElementTree as ET

from openpyxl.utils.cell import (
    column_index_from_string,
    coordinate_from_string,
    get_column_letter,
    range_boundaries,
)
from openpyxl.utils.exceptions import CellCoordinatesException

from fdm.models import ImageDocument
from fdm.settings import (
    RawRecordDataSource,
    RawRecordExportDirection,
    RawRecordExportRule,
    RawRecordMeasurementFilter,
    RawRecordTemplate,
    SUPPORTED_RAW_RECORD_TEMPLATE_SUFFIXES,
    resolve_resource_relative_path,
)

SPREADSHEET_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
OFFICE_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
PACKAGE_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
CONTENT_TYPES_NS = "http://schemas.openxmlformats.org/package/2006/content-types"
MARKUP_COMPAT_NS = "http://schemas.openxmlformats.org/markup-compatibility/2006"
X14_NS = "http://schemas.microsoft.com/office/spreadsheetml/2009/9/main"
X14AC_NS = "http://schemas.microsoft.com/office/spreadsheetml/2009/9/ac"
X15_NS = "http://schemas.microsoft.com/office/spreadsheetml/2010/11/main"
X15AC_NS = "http://schemas.microsoft.com/office/spreadsheetml/2010/11/ac"
XML_NS = "http://www.w3.org/XML/1998/namespace"

ET.register_namespace("", SPREADSHEET_NS)
ET.register_namespace("r", OFFICE_REL_NS)
ET.register_namespace("mc", MARKUP_COMPAT_NS)
ET.register_namespace("x14", X14_NS)
ET.register_namespace("x14ac", X14AC_NS)
ET.register_namespace("x15", X15_NS)
ET.register_namespace("x15ac", X15AC_NS)
ET.register_namespace("pkgrel", PACKAGE_REL_NS)
ET.register_namespace("ct", CONTENT_TYPES_NS)

RAW_RECORD_FIELD_NAMES = [
    "纤维种类",
    "类型",
    "结果",
    "单位",
    "标尺信息",
    "模式",
    "状态",
    "置信度",
    "起点X(px)",
    "起点Y(px)",
    "终点X(px)",
    "终点Y(px)",
    "像素直径(px)",
    "多边形点数",
    "多边形顶点JSON",
    "折线点数",
    "折线顶点JSON",
    "计数点X(px)",
    "计数点Y(px)",
    "面积(px²)",
    "孔洞面积",
    "创建时间",
    "图片路径",
    "图片编号",
    "测量记录ID",
    "纤维种类ID",
    "纤维类别序号",
    "纤维种类编号",
    "纤维种类名称",
    "纤维种类颜色",
]

MEASUREMENT_KIND_ROW_LABELS = {
    RawRecordMeasurementFilter.LINE: "线段",
    RawRecordMeasurementFilter.AREA: "面积",
    RawRecordMeasurementFilter.POLYLINE: "折线",
    RawRecordMeasurementFilter.COUNT: "计数点",
}


class RawRecordTemplateExportError(RuntimeError):
    pass


@dataclass(frozen=True, slots=True)
class _CellWrite:
    coordinate: str
    value: object


@dataclass(frozen=True, slots=True)
class _RuleWritePlan:
    rule: RawRecordExportRule
    values: list[object]


def raw_record_output_suffix(template_path: str | Path) -> str:
    suffix = Path(str(template_path)).suffix.lower()
    if suffix in {".xlsm", ".xltm"}:
        return ".xlsm"
    return ".xlsx"


def raw_record_output_path(output_path: str | Path, template_path: str | Path) -> Path:
    return Path(output_path).with_suffix(raw_record_output_suffix(template_path))


def write_raw_record_template(
    template: RawRecordTemplate,
    output_path: str | Path,
    *,
    documents: list[ImageDocument],
    measurement_rows: list[dict[str, object]],
) -> Path:
    template_path = resolve_resource_relative_path(template.path)
    if not template_path.exists():
        raise FileNotFoundError(str(template_path))
    if template_path.suffix.lower() not in SUPPORTED_RAW_RECORD_TEMPLATE_SUFFIXES:
        raise RawRecordTemplateExportError(f"不支持的原始记录模板格式: {template_path.suffix or '(无扩展名)'}")

    target_path = raw_record_output_path(output_path, template_path)
    if _same_path(template_path, target_path):
        raise RawRecordTemplateExportError("导出目标不能覆盖原始记录模板文件。")
    target_path.parent.mkdir(parents=True, exist_ok=True)

    try:
        with zipfile.ZipFile(template_path, "r") as archive:
            entries = [(info, archive.read(info.filename)) for info in archive.infolist()]
    except zipfile.BadZipFile as exc:
        raise RawRecordTemplateExportError("原始记录模板不是有效的 OOXML Excel 文件，请先另存为 .xlsx/.xlsm/.xltx/.xltm。") from exc

    try:
        entry_data = {info.filename: data for info, data in entries}
        sheet_paths = _workbook_sheet_paths(entry_data)
        rule_plans = _build_rule_plans(template.rules, documents=documents, measurement_rows=measurement_rows)
        sheet_write_plans = _group_rule_plans_by_sheet(rule_plans, sheet_paths)

        modified_entries = dict(entry_data)
        for sheet_name, plans in sheet_write_plans.items():
            sheet_path = sheet_paths[sheet_name]
            modified_entries[sheet_path] = _updated_sheet_xml(modified_entries[sheet_path], plans, sheet_name=sheet_name)
        modified_entries["[Content_Types].xml"] = _updated_content_types_xml(
            modified_entries["[Content_Types].xml"],
            input_suffix=template_path.suffix.lower(),
            output_suffix=target_path.suffix.lower(),
        )
        _remove_calc_chain(modified_entries)
    except KeyError as exc:
        raise RawRecordTemplateExportError("原始记录模板缺少必要的 OOXML 文件。") from exc
    except ET.ParseError as exc:
        raise RawRecordTemplateExportError("原始记录模板 XML 结构无法解析。") from exc

    temp_name = ""
    try:
        with tempfile.NamedTemporaryFile(
            "wb",
            delete=False,
            dir=target_path.parent,
            prefix=f".{target_path.stem}.",
            suffix=target_path.suffix,
        ) as temp_file:
            temp_name = temp_file.name
        with zipfile.ZipFile(temp_name, "w") as output_archive:
            for info, _data in entries:
                if info.filename not in modified_entries:
                    continue
                output_archive.writestr(info, modified_entries[info.filename])
        _validate_ooxml_package(
            temp_name,
            original_entries=entry_data,
            source_suffix=template_path.suffix.lower(),
            output_suffix=target_path.suffix.lower(),
        )
        Path(temp_name).replace(target_path)
    except Exception:
        if temp_name:
            try:
                Path(temp_name).unlink(missing_ok=True)
            except OSError:
                pass
        raise
    return target_path


def _same_path(left: Path, right: Path) -> bool:
    try:
        return left.resolve() == right.resolve()
    except OSError:
        return os.path.abspath(left) == os.path.abspath(right)


def _build_rule_plans(
    rules: list[RawRecordExportRule],
    *,
    documents: list[ImageDocument],
    measurement_rows: list[dict[str, object]],
) -> list[_RuleWritePlan]:
    plans: list[_RuleWritePlan] = []
    for raw_rule in rules:
        rule = raw_rule.normalized_copy()
        values = _values_for_rule(rule, documents=documents, measurement_rows=measurement_rows)
        plans.append(_RuleWritePlan(rule=rule, values=values))
    return plans


def _values_for_rule(
    rule: RawRecordExportRule,
    *,
    documents: list[ImageDocument],
    measurement_rows: list[dict[str, object]],
) -> list[object]:
    if rule.data_source == RawRecordDataSource.DIAMETER_RESULT:
        return [
            measurement.display_value()
            for document in documents
            for measurement in document.measurements
            if measurement.measurement_kind == "line"
        ]
    if rule.data_source == RawRecordDataSource.AREA_RESULT:
        return [
            measurement.display_value()
            for document in documents
            for measurement in document.measurements
            if measurement.measurement_kind == "area"
        ]
    if rule.data_source == RawRecordDataSource.UNIQUE_FIELD_RANGE:
        return _unique_field_values(rule, measurement_rows)

    field_name = rule.field_name or "结果"
    return [
        row.get(field_name)
        for row in measurement_rows
        if _row_matches_measurement_filter(row, rule.measurement_filter)
    ]


def _unique_field_values(rule: RawRecordExportRule, measurement_rows: list[dict[str, object]]) -> list[object]:
    field_name = rule.field_name or "纤维种类名称"
    values: list[object] = []
    seen_labels: set[str] = set()
    for row in measurement_rows:
        if not _row_matches_measurement_filter(row, rule.measurement_filter):
            continue
        label = str(row.get("纤维种类名称", "") or "").strip()
        if not label or label in seen_labels:
            continue
        seen_labels.add(label)
        values.append(row.get(field_name))
    return values


def _row_matches_measurement_filter(row: dict[str, object], measurement_filter: str) -> bool:
    if measurement_filter == RawRecordMeasurementFilter.ALL:
        return True
    return str(row.get("类型", "")) == MEASUREMENT_KIND_ROW_LABELS.get(measurement_filter, "")


def _group_rule_plans_by_sheet(
    rule_plans: list[_RuleWritePlan],
    sheet_paths: dict[str, str],
) -> dict[str, list[_RuleWritePlan]]:
    grouped: dict[str, list[_RuleWritePlan]] = defaultdict(list)
    occupied: dict[tuple[str, str], RawRecordExportRule] = {}
    for plan in rule_plans:
        sheet_name = plan.rule.sheet_name.strip()
        if sheet_name not in sheet_paths:
            raise RawRecordTemplateExportError(f"原始记录模板中找不到工作表: {sheet_name}")
        coordinates = _occupied_coordinates_for_rule(plan.rule, len(plan.values))
        for coordinate in coordinates:
            key = (sheet_name, coordinate)
            if key in occupied:
                raise RawRecordTemplateExportError(f"原始记录导出规则目标单元格重叠: {sheet_name}!{coordinate}")
            occupied[key] = plan.rule
        grouped[sheet_name].append(plan)
    return dict(grouped)


def _target_coordinates(start_cell: str, direction: str, value_count: int) -> list[str]:
    row, column = _parse_cell_coordinate(start_cell)
    coordinates: list[str] = []
    for offset in range(value_count):
        target_row = row + offset if direction == RawRecordExportDirection.VERTICAL else row
        target_column = column if direction == RawRecordExportDirection.VERTICAL else column + offset
        coordinates.append(f"{get_column_letter(target_column)}{target_row}")
    return coordinates


def _target_coordinates_for_rule(rule: RawRecordExportRule, value_count: int) -> list[str]:
    if rule.data_source == RawRecordDataSource.UNIQUE_FIELD_RANGE:
        return _range_target_coordinates(rule.start_cell, rule.end_cell, rule.direction, value_count)
    return _target_coordinates(rule.start_cell, rule.direction, value_count)


def _occupied_coordinates_for_rule(rule: RawRecordExportRule, value_count: int) -> list[str]:
    if rule.data_source == RawRecordDataSource.UNIQUE_FIELD_RANGE:
        return _range_target_coordinates(rule.start_cell, rule.end_cell, rule.direction, None)
    return _target_coordinates(rule.start_cell, rule.direction, value_count)


def _range_target_coordinates(start_cell: str, end_cell: str, direction: str, value_count: int | None) -> list[str]:
    if not str(end_cell or "").strip():
        raise RawRecordTemplateExportError("去重字段范围规则需要填写结束单元格。")
    start_row, start_column = _parse_cell_coordinate(start_cell)
    end_row, end_column = _parse_cell_coordinate(end_cell)
    if direction == RawRecordExportDirection.HORIZONTAL:
        if start_row != end_row or end_column < start_column:
            raise RawRecordTemplateExportError(f"横向去重字段范围无效: {start_cell}:{end_cell}")
        capacity = end_column - start_column + 1
        target_count = capacity if value_count is None else min(value_count, capacity)
        return [
            f"{get_column_letter(start_column + offset)}{start_row}"
            for offset in range(target_count)
        ]
    if start_column != end_column or end_row < start_row:
        raise RawRecordTemplateExportError(f"纵向去重字段范围无效: {start_cell}:{end_cell}")
    capacity = end_row - start_row + 1
    target_count = capacity if value_count is None else min(value_count, capacity)
    return [
        f"{get_column_letter(start_column)}{start_row + offset}"
        for offset in range(target_count)
    ]


def _parse_cell_coordinate(coordinate: str) -> tuple[int, int]:
    try:
        column_letter, row = coordinate_from_string(str(coordinate).strip().upper())
        column = column_index_from_string(column_letter)
    except (CellCoordinatesException, ValueError) as exc:
        raise RawRecordTemplateExportError(f"原始记录导出规则单元格无效: {coordinate}") from exc
    if row < 1 or column < 1:
        raise RawRecordTemplateExportError(f"原始记录导出规则单元格无效: {coordinate}")
    return row, column


def _workbook_sheet_paths(entry_data: dict[str, bytes]) -> dict[str, str]:
    try:
        workbook = ET.fromstring(entry_data["xl/workbook.xml"])
        rels = ET.fromstring(entry_data["xl/_rels/workbook.xml.rels"])
    except KeyError as exc:
        raise RawRecordTemplateExportError("原始记录模板缺少 Excel 工作簿结构。") from exc
    relationship_targets = {
        relationship.attrib.get("Id", ""): relationship.attrib.get("Target", "")
        for relationship in rels.findall(f"{{{PACKAGE_REL_NS}}}Relationship")
    }
    sheet_paths: dict[str, str] = {}
    for sheet in workbook.findall(f".//{{{SPREADSHEET_NS}}}sheet"):
        sheet_name = sheet.attrib.get("name", "")
        relationship_id = sheet.attrib.get(f"{{{OFFICE_REL_NS}}}id", "")
        target = relationship_targets.get(relationship_id, "")
        if not sheet_name or not target:
            continue
        sheet_paths[sheet_name] = _normalize_workbook_target_path(target)
    if not sheet_paths:
        raise RawRecordTemplateExportError("原始记录模板中没有可写入的工作表。")
    return sheet_paths


def _normalize_workbook_target_path(target: str) -> str:
    token = target.lstrip("/")
    if token.startswith("xl/"):
        return token
    return f"xl/{token}"


def _register_namespaces_from_xml(xml_bytes: bytes) -> dict[str, str]:
    fallback_namespaces: dict[str, str] = {
        "": SPREADSHEET_NS,
        "r": OFFICE_REL_NS,
        "mc": MARKUP_COMPAT_NS,
        "x14": X14_NS,
        "x14ac": X14AC_NS,
        "x15": X15_NS,
        "x15ac": X15AC_NS,
    }
    namespaces: dict[str, str] = {}
    try:
        for _event, namespace in ET.iterparse(BytesIO(xml_bytes), events=("start-ns",)):
            prefix, uri = namespace
            namespaces[prefix or ""] = uri
    except ET.ParseError:
        raise
    for prefix, uri in {**fallback_namespaces, **namespaces}.items():
        if prefix == "xml":
            continue
        try:
            ET.register_namespace(prefix, uri)
        except ValueError:
            continue
    return namespaces


def _preserve_compatibility_namespace_declarations(xml_bytes: bytes, namespaces: dict[str, str]) -> bytes:
    text = xml_bytes.decode("utf-8")
    missing_declarations = [
        (prefix, namespaces[prefix])
        for prefix in sorted(_compatibility_prefixes(text))
        if prefix in namespaces and f"xmlns:{prefix}=" not in text
    ]
    if not missing_declarations:
        return xml_bytes

    declaration_end = text.find("?>")
    search_start = declaration_end + 2 if declaration_end >= 0 else 0
    root_start = text.find("<", search_start)
    root_end = text.find(">", root_start)
    if root_start < 0 or root_end < 0:
        return xml_bytes
    insertion = "".join(f' xmlns:{prefix}="{uri}"' for prefix, uri in missing_declarations)
    return f"{text[:root_end]}{insertion}{text[root_end:]}".encode("utf-8")


def _updated_sheet_xml(sheet_xml: bytes, rule_plans: list[_RuleWritePlan], *, sheet_name: str) -> bytes:
    original_namespaces = _register_namespaces_from_xml(sheet_xml)
    root = ET.fromstring(sheet_xml)
    sheet_data = root.find(f"{{{SPREADSHEET_NS}}}sheetData")
    if sheet_data is None:
        sheet_data = ET.SubElement(root, f"{{{SPREADSHEET_NS}}}sheetData")

    row_map = _row_map(sheet_data)
    merged_ranges = _merged_ranges(root)
    for plan in rule_plans:
        start_style = _existing_cell_style(row_map, plan.rule.start_cell)
        writes = [
            _CellWrite(coordinate=coordinate, value=value)
            for coordinate, value in zip(
                _target_coordinates_for_rule(plan.rule, len(plan.values)),
                plan.values,
            )
        ]
        for write in writes:
            _validate_merge_target(write.coordinate, merged_ranges, sheet_name=sheet_name)
            cell = _get_or_create_cell(sheet_data, row_map, write.coordinate, inherited_style=start_style)
            _write_cell_value(cell, write.value)

    _sort_sheet_rows_and_cells(sheet_data)
    _refresh_dimension(root)
    updated_xml = ET.tostring(root, encoding="utf-8", xml_declaration=True)
    return _preserve_compatibility_namespace_declarations(updated_xml, original_namespaces)


def _row_map(sheet_data: ET.Element) -> dict[int, ET.Element]:
    rows: dict[int, ET.Element] = {}
    for row in sheet_data.findall(f"{{{SPREADSHEET_NS}}}row"):
        try:
            row_number = int(row.attrib.get("r", "0"))
        except ValueError:
            continue
        if row_number > 0:
            rows[row_number] = row
    return rows


def _existing_cell_style(row_map: dict[int, ET.Element], coordinate: str) -> str | None:
    row_number, _column = _parse_cell_coordinate(coordinate)
    row = row_map.get(row_number)
    if row is None:
        return None
    cell = _find_cell(row, coordinate)
    return cell.attrib.get("s") if cell is not None else None


def _get_or_create_cell(
    sheet_data: ET.Element,
    row_map: dict[int, ET.Element],
    coordinate: str,
    *,
    inherited_style: str | None,
) -> ET.Element:
    row_number, _column = _parse_cell_coordinate(coordinate)
    row = row_map.get(row_number)
    if row is None:
        row = ET.SubElement(sheet_data, f"{{{SPREADSHEET_NS}}}row", {"r": str(row_number)})
        row_map[row_number] = row
    cell = _find_cell(row, coordinate)
    if cell is not None:
        return cell
    attributes = {"r": coordinate}
    if inherited_style:
        attributes["s"] = inherited_style
    return ET.SubElement(row, f"{{{SPREADSHEET_NS}}}c", attributes)


def _find_cell(row: ET.Element, coordinate: str) -> ET.Element | None:
    for cell in row.findall(f"{{{SPREADSHEET_NS}}}c"):
        if cell.attrib.get("r", "").upper() == coordinate.upper():
            return cell
    return None


def _write_cell_value(cell: ET.Element, value: object) -> None:
    _clear_cell_value(cell)
    if _is_empty_value(value):
        return
    if isinstance(value, (int, float)) and not isinstance(value, bool) and math.isfinite(float(value)):
        cell.attrib.pop("t", None)
        value_node = ET.SubElement(cell, f"{{{SPREADSHEET_NS}}}v")
        value_node.text = _number_text(value)
        return
    if isinstance(value, bool):
        cell.set("t", "b")
        value_node = ET.SubElement(cell, f"{{{SPREADSHEET_NS}}}v")
        value_node.text = "1" if value else "0"
        return
    cell.set("t", "inlineStr")
    inline_string = ET.SubElement(cell, f"{{{SPREADSHEET_NS}}}is")
    text_node = ET.SubElement(inline_string, f"{{{SPREADSHEET_NS}}}t")
    text = str(value)
    if text != text.strip():
        text_node.set(f"{{{XML_NS}}}space", "preserve")
    text_node.text = text


def _clear_cell_value(cell: ET.Element) -> None:
    cell.attrib.pop("t", None)
    for child in list(cell):
        if child.tag in {
            f"{{{SPREADSHEET_NS}}}f",
            f"{{{SPREADSHEET_NS}}}v",
            f"{{{SPREADSHEET_NS}}}is",
        }:
            cell.remove(child)


def _is_empty_value(value: object) -> bool:
    if value is None:
        return True
    if isinstance(value, float) and math.isnan(value):
        return True
    return False


def _number_text(value: int | float) -> str:
    if isinstance(value, int):
        return str(value)
    if float(value).is_integer():
        return str(int(value))
    return format(float(value), ".15g")


def _merged_ranges(root: ET.Element) -> list[tuple[int, int, int, int]]:
    ranges: list[tuple[int, int, int, int]] = []
    for merge_cell in root.findall(f".//{{{SPREADSHEET_NS}}}mergeCell"):
        ref = merge_cell.attrib.get("ref", "")
        try:
            min_col, min_row, max_col, max_row = range_boundaries(ref)
        except ValueError:
            continue
        ranges.append((min_col, min_row, max_col, max_row))
    return ranges


def _validate_merge_target(
    coordinate: str,
    merged_ranges: list[tuple[int, int, int, int]],
    *,
    sheet_name: str,
) -> None:
    row, column = _parse_cell_coordinate(coordinate)
    for min_col, min_row, max_col, max_row in merged_ranges:
        if min_col <= column <= max_col and min_row <= row <= max_row:
            if column == min_col and row == min_row:
                return
            raise RawRecordTemplateExportError(f"不能写入合并单元格的非左上角位置: {sheet_name}!{coordinate}")


def _sort_sheet_rows_and_cells(sheet_data: ET.Element) -> None:
    rows = list(sheet_data)
    for row in rows:
        sheet_data.remove(row)
    for row in sorted(rows, key=_row_sort_key):
        _sort_row_cells(row)
        sheet_data.append(row)


def _row_sort_key(row: ET.Element) -> int:
    try:
        return int(row.attrib.get("r", "0"))
    except ValueError:
        return 0


def _sort_row_cells(row: ET.Element) -> None:
    cells = [child for child in list(row) if child.tag == f"{{{SPREADSHEET_NS}}}c"]
    others = [child for child in list(row) if child.tag != f"{{{SPREADSHEET_NS}}}c"]
    for child in list(row):
        row.remove(child)
    for cell in sorted(cells, key=_cell_sort_key):
        row.append(cell)
    for child in others:
        row.append(child)


def _cell_sort_key(cell: ET.Element) -> int:
    coordinate = cell.attrib.get("r", "A1")
    try:
        _row, column = _parse_cell_coordinate(coordinate)
    except RawRecordTemplateExportError:
        return 0
    return column


def _refresh_dimension(root: ET.Element) -> None:
    min_row: int | None = None
    min_column: int | None = None
    max_row = 0
    max_column = 0
    for cell in root.findall(f".//{{{SPREADSHEET_NS}}}c"):
        coordinate = cell.attrib.get("r", "")
        if not coordinate:
            continue
        try:
            row, column = _parse_cell_coordinate(coordinate)
        except RawRecordTemplateExportError:
            continue
        min_row = row if min_row is None else min(min_row, row)
        min_column = column if min_column is None else min(min_column, column)
        max_row = max(max_row, row)
        max_column = max(max_column, column)
    dimension = root.find(f"{{{SPREADSHEET_NS}}}dimension")
    if dimension is None:
        dimension = ET.Element(f"{{{SPREADSHEET_NS}}}dimension")
        root.insert(0, dimension)
    if min_row is None or min_column is None:
        dimension.set("ref", "A1")
        return
    start = f"{get_column_letter(min_column)}{min_row}"
    end = f"{get_column_letter(max_column)}{max_row}"
    dimension.set("ref", start if start == end else f"{start}:{end}")


def _updated_content_types_xml(content_types_xml: bytes, *, input_suffix: str, output_suffix: str) -> bytes:
    updated = content_types_xml
    if input_suffix == ".xltm" and output_suffix == ".xlsm":
        updated = updated.replace(
            b"application/vnd.ms-excel.template.macroEnabled.main+xml",
            b"application/vnd.ms-excel.sheet.macroEnabled.main+xml",
        )
    elif input_suffix == ".xltx" and output_suffix == ".xlsx":
        updated = updated.replace(
            b"application/vnd.openxmlformats-officedocument.spreadsheetml.template.main+xml",
            b"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml",
        )
    return updated


def _remove_calc_chain(entries: dict[str, bytes]) -> None:
    if "xl/calcChain.xml" not in entries:
        return
    entries.pop("xl/calcChain.xml", None)
    if "xl/_rels/workbook.xml.rels" in entries:
        entries["xl/_rels/workbook.xml.rels"] = _remove_calc_chain_relationship(
            entries["xl/_rels/workbook.xml.rels"]
        )
    if "[Content_Types].xml" in entries:
        entries["[Content_Types].xml"] = _remove_calc_chain_content_type(
            entries["[Content_Types].xml"]
        )


def _remove_calc_chain_relationship(rels_xml: bytes) -> bytes:
    text = rels_xml.decode("utf-8")
    pattern = re.compile(r"\s*<Relationship\b[^>]*/>", re.IGNORECASE)

    def replacement(match: re.Match[str]) -> str:
        fragment = match.group(0)
        if "calcChain" in fragment and "Relationship" in fragment:
            return ""
        return fragment

    return pattern.sub(replacement, text).encode("utf-8")


def _remove_calc_chain_content_type(content_types_xml: bytes) -> bytes:
    text = content_types_xml.decode("utf-8")
    pattern = re.compile(
        r"\s*<Override\b(?=[^>]*\bPartName=(['\"])/xl/calcChain\.xml\1)[^>]*/>",
        re.IGNORECASE,
    )
    return pattern.sub("", text).encode("utf-8")


def _validate_ooxml_package(
    package_path: str | Path,
    *,
    original_entries: dict[str, bytes],
    source_suffix: str,
    output_suffix: str,
) -> None:
    try:
        with zipfile.ZipFile(package_path, "r") as archive:
            entries = {
                name: archive.read(name)
                for name in archive.namelist()
                if not name.endswith("/")
            }
    except zipfile.BadZipFile as exc:
        raise RawRecordTemplateExportError("原始记录模板导出结果不是有效的 OOXML 文件。") from exc

    _validate_xml_parts(entries)
    _validate_relationship_targets(entries)
    _validate_macro_part_preserved(entries, original_entries, source_suffix=source_suffix, output_suffix=output_suffix)


def _validate_xml_parts(entries: dict[str, bytes]) -> None:
    for name, data in entries.items():
        if not name.endswith(".xml"):
            continue
        try:
            ET.fromstring(data)
        except ET.ParseError as exc:
            raise RawRecordTemplateExportError(f"原始记录模板导出结果 XML 无法解析: {name}") from exc
        missing_prefixes = _missing_compatibility_prefixes(data)
        if missing_prefixes:
            missing_text = ", ".join(missing_prefixes)
            raise RawRecordTemplateExportError(f"原始记录模板导出结果缺少兼容性命名空间声明: {name} ({missing_text})")


def _missing_compatibility_prefixes(xml_bytes: bytes) -> list[str]:
    text = xml_bytes.decode("utf-8", errors="replace")
    declarations = _namespace_declarations(text)
    return sorted(prefix for prefix in _compatibility_prefixes(text) if prefix not in declarations)


def _namespace_declarations(xml_text: str) -> dict[str, str]:
    return {
        match.group(1): match.group(3)
        for match in re.finditer(r"\bxmlns:([A-Za-z_][\w.-]*)=(['\"])(.*?)\2", xml_text)
    }


def _compatibility_prefixes(xml_text: str) -> set[str]:
    prefixes: set[str] = set()
    pattern = re.compile(r"\b(?:[A-Za-z_][\w.-]*:)?(?:Ignorable|Requires|ProcessContent)=(['\"])(.*?)\1")
    for match in pattern.finditer(xml_text):
        for token in str(match.group(2) or "").split():
            if not token:
                continue
            prefix = token.split(":", 1)[0]
            if prefix and prefix != "xml":
                prefixes.add(prefix)
    return prefixes


def _validate_relationship_targets(entries: dict[str, bytes]) -> None:
    for rel_name, rel_xml in entries.items():
        if not rel_name.endswith(".rels"):
            continue
        try:
            root = ET.fromstring(rel_xml)
        except ET.ParseError as exc:
            raise RawRecordTemplateExportError(f"原始记录模板导出结果关系文件无法解析: {rel_name}") from exc
        base_path = _relationship_base_path(rel_name)
        for relationship in root.findall(f"{{{PACKAGE_REL_NS}}}Relationship"):
            target_mode = relationship.attrib.get("TargetMode", "")
            target = relationship.attrib.get("Target", "")
            if not target or target_mode == "External" or _is_external_relationship_target(target):
                continue
            target_path = _resolve_relationship_target(base_path, target)
            if target_path not in entries:
                raise RawRecordTemplateExportError(f"原始记录模板导出结果关系目标缺失: {rel_name} -> {target}")


def _relationship_base_path(rel_name: str) -> str:
    if rel_name == "_rels/.rels":
        return ""
    if "/_rels/" in rel_name:
        prefix, _tail = rel_name.split("/_rels/", 1)
        return prefix
    return posixpath.dirname(rel_name)


def _resolve_relationship_target(base_path: str, target: str) -> str:
    if target.startswith("/"):
        return posixpath.normpath(target.lstrip("/"))
    return posixpath.normpath(posixpath.join(base_path, target))


def _is_external_relationship_target(target: str) -> bool:
    token = target.strip().lower()
    return bool(
        "://" in token
        or token.startswith("mailto:")
        or token.startswith("file:")
        or token.startswith("urn:")
    )


def _validate_macro_part_preserved(
    entries: dict[str, bytes],
    original_entries: dict[str, bytes],
    *,
    source_suffix: str,
    output_suffix: str,
) -> None:
    if source_suffix not in {".xlsm", ".xltm"} or output_suffix != ".xlsm":
        return
    original_macro = original_entries.get("xl/vbaProject.bin")
    if original_macro is None:
        return
    if entries.get("xl/vbaProject.bin") != original_macro:
        raise RawRecordTemplateExportError("原始记录模板导出结果中的宏工程已变化。")
