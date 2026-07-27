"""Independent exports for :mod:`fdm.analysis_artifacts`.

These exports intentionally do not reuse the measurement/raw-record workbook
pipeline.  Analysis artifacts have a different lifecycle, schema and invalid
state, so mixing them into measurement columns would make both formats
ambiguous.
"""

from __future__ import annotations

from collections.abc import Iterable, Mapping
import csv
from dataclasses import dataclass
from datetime import datetime, timezone
from enum import StrEnum
import hashlib
from io import BytesIO, StringIO
import json
from pathlib import Path, PurePosixPath
import re
from typing import TypeAlias
import unicodedata
import zipfile

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from fdm.analysis_artifacts import (
    AnalysisArtifact,
    AnalysisArtifactStatus,
    AnalysisCurve,
    AnalysisObjectKind,
    AnalysisTable,
    JsonScalar,
)
from fdm.atomic_io import atomic_write_bytes
from fdm.services.analysis_batch import (
    AnalysisBatchItemResult,
    AnalysisBatchResult,
)
from fdm.services.analysis_profiles import (
    ANALYSIS_OUTPUT_FIELDS_PARAMETER,
    analysis_output_field_schema,
)


class AnalysisExportFailureCode(StrEnum):
    EMPTY_SELECTION = "empty_selection"
    INVALID_TARGET = "invalid_target"
    TABLE_NOT_FOUND = "table_not_found"
    CURVE_NOT_FOUND = "curve_not_found"
    ASSET_NOT_FOUND = "asset_not_found"
    ASSET_HASH_MISMATCH = "asset_hash_mismatch"
    ENCODE_FAILED = "encode_failed"
    VERIFY_FAILED = "verify_failed"
    ATOMIC_COMMIT_FAILED = "atomic_commit_failed"


@dataclass(frozen=True, slots=True)
class AnalysisExportResult:
    success: bool
    path: Path | None
    artifact_count: int
    sheet_count: int
    message: str
    failure_code: AnalysisExportFailureCode | None = None

    @classmethod
    def succeeded(
        cls,
        path: Path,
        *,
        artifact_count: int,
        sheet_count: int,
    ) -> "AnalysisExportResult":
        return cls(
            success=True,
            path=path,
            artifact_count=artifact_count,
            sheet_count=sheet_count,
            message=f"已导出到 {path}",
        )

    @classmethod
    def failed(
        cls,
        code: AnalysisExportFailureCode,
        message: str,
        *,
        path: Path | None = None,
        artifact_count: int = 0,
    ) -> "AnalysisExportResult":
        return cls(
            success=False,
            path=path,
            artifact_count=artifact_count,
            sheet_count=0,
            message=str(message),
            failure_code=code,
        )


@dataclass(frozen=True, slots=True)
class AnalysisBatchExportRow:
    """One immutable, auditable row derived from a batch item result."""

    item_id: str
    display_name: str
    document_id: str
    document_name: str
    roi_id: str | None
    roi_name: str | None
    success: bool
    step_index: int
    step_count: int
    tool_id: str
    tool_name: str
    algorithm_version: str
    scalar_report: tuple[tuple[str, JsonScalar], ...]
    array_summaries: tuple[str, ...]
    error_type: str | None = None
    error_message: str | None = None

    @property
    def scope(self) -> str:
        return "ROI" if self.roi_id is not None else "整张图片"

    @property
    def step_label(self) -> str:
        if self.step_index < 1:
            return "—"
        return f"{self.step_index}/{self.step_count}"


@dataclass(frozen=True, slots=True)
class AnalysisBatchWorkbookExportResult:
    success: bool
    path: Path | None
    item_count: int
    success_count: int
    failure_count: int
    sheet_count: int
    message: str
    failure_code: AnalysisExportFailureCode | None = None

    @classmethod
    def succeeded(
        cls,
        path: Path,
        *,
        item_count: int,
        success_count: int,
        failure_count: int,
        sheet_count: int,
    ) -> "AnalysisBatchWorkbookExportResult":
        return cls(
            success=True,
            path=path,
            item_count=item_count,
            success_count=success_count,
            failure_count=failure_count,
            sheet_count=sheet_count,
            message=f"已导出批量分析工作簿到 {path}",
        )

    @classmethod
    def failed(
        cls,
        code: AnalysisExportFailureCode,
        message: str,
        *,
        path: Path | None = None,
        item_count: int = 0,
        success_count: int = 0,
        failure_count: int = 0,
    ) -> "AnalysisBatchWorkbookExportResult":
        return cls(
            success=False,
            path=path,
            item_count=item_count,
            success_count=success_count,
            failure_count=failure_count,
            sheet_count=0,
            message=str(message),
            failure_code=code,
        )


ArtifactNameMap: TypeAlias = Mapping[str, str] | None

_TOOL_NAMES = {
    "fdm.shape": "形状测量",
    "fdm.shape_analysis": "形状测量",
    "fdm.intensity": "灰度与颜色统计",
    "fdm.intensity_statistics": "灰度与颜色统计",
    "fdm.histogram": "直方图",
    "fdm.fft_power_spectrum": "FFT 功率谱",
    "fdm.profile": "强度剖面",
    "fdm.intensity_profile": "强度剖面",
    "fdm.particles": "粒子分析",
    "fdm.particle_analysis": "粒子分析",
    "fdm.maxima": "极值检测",
    "fdm.maxima_detection": "极值检测",
}
_INVALID_SHEET_CHARS = re.compile(r"[\[\]:*?/\\]")
_FORMULA_PREFIXES = ("=", "+", "-", "@")


class AnalysisExportService:
    """Create auditable analysis workbooks and individual table CSV files."""

    def export_workbook(
        self,
        artifacts: Iterable[AnalysisArtifact],
        target_path: str | Path,
        *,
        document_names: ArtifactNameMap = None,
        roi_names: ArtifactNameMap = None,
        measurement_names: ArtifactNameMap = None,
        tool_names: ArtifactNameMap = None,
    ) -> AnalysisExportResult:
        frozen = tuple(artifacts)
        try:
            target = _normalized_target(target_path, ".xlsx")
        except (TypeError, ValueError, OSError) as exc:
            return AnalysisExportResult.failed(
                AnalysisExportFailureCode.INVALID_TARGET,
                f"分析工作簿路径无效：{_error_message(exc)}",
            )
        if not frozen:
            return AnalysisExportResult.failed(
                AnalysisExportFailureCode.EMPTY_SELECTION,
                "没有可导出的分析结果。",
                path=target,
            )
        if any(not isinstance(item, AnalysisArtifact) for item in frozen):
            return AnalysisExportResult.failed(
                AnalysisExportFailureCode.ENCODE_FAILED,
                "分析结果列表包含不支持的对象。",
                path=target,
            )
        try:
            workbook = _build_workbook(
                frozen,
                document_names=document_names,
                roi_names=roi_names,
                measurement_names=measurement_names,
                tool_names=tool_names,
            )
            buffer = BytesIO()
            workbook.save(buffer)
            payload = buffer.getvalue()
        except Exception as exc:
            return AnalysisExportResult.failed(
                AnalysisExportFailureCode.ENCODE_FAILED,
                f"无法生成分析工作簿：{_error_message(exc)}",
                path=target,
                artifact_count=len(frozen),
            )

        try:
            _verify_workbook(payload)
        except Exception as exc:
            return AnalysisExportResult.failed(
                AnalysisExportFailureCode.VERIFY_FAILED,
                f"分析工作簿校验失败：{_error_message(exc)}",
                path=target,
                artifact_count=len(frozen),
            )
        try:
            atomic_write_bytes(target, payload)
        except Exception as exc:
            return AnalysisExportResult.failed(
                AnalysisExportFailureCode.ATOMIC_COMMIT_FAILED,
                f"无法写入分析工作簿：{_error_message(exc)}",
                path=target,
                artifact_count=len(frozen),
            )
        if not target.is_file() or target.stat().st_size <= 0:
            return AnalysisExportResult.failed(
                AnalysisExportFailureCode.VERIFY_FAILED,
                "分析工作簿写入后为空。",
                path=target,
                artifact_count=len(frozen),
            )
        return AnalysisExportResult.succeeded(
            target,
            artifact_count=len(frozen),
            sheet_count=len(workbook.sheetnames),
        )

    def export_batch_workbook(
        self,
        result: AnalysisBatchResult,
        target_path: str | Path,
        *,
        document_names: ArtifactNameMap = None,
        roi_names: ArtifactNameMap = None,
    ) -> AnalysisBatchWorkbookExportResult:
        try:
            target = _normalized_target(target_path, ".xlsx")
        except (TypeError, ValueError, OSError) as exc:
            return AnalysisBatchWorkbookExportResult.failed(
                AnalysisExportFailureCode.INVALID_TARGET,
                f"批量分析工作簿路径无效：{_error_message(exc)}",
            )
        if not isinstance(result, AnalysisBatchResult):
            return AnalysisBatchWorkbookExportResult.failed(
                AnalysisExportFailureCode.ENCODE_FAILED,
                "批量分析结果对象无效。",
                path=target,
            )
        item_count = len(result.item_results)
        success_count = result.success_count
        failure_count = result.failure_count
        try:
            rows = build_analysis_batch_export_rows(
                result,
                document_names=document_names,
                roi_names=roi_names,
            )
            workbook = _build_analysis_batch_workbook(result, rows)
            buffer = BytesIO()
            workbook.save(buffer)
            payload = buffer.getvalue()
        except Exception as exc:
            return AnalysisBatchWorkbookExportResult.failed(
                AnalysisExportFailureCode.ENCODE_FAILED,
                f"无法生成批量分析工作簿：{_error_message(exc)}",
                path=target,
                item_count=item_count,
                success_count=success_count,
                failure_count=failure_count,
            )
        try:
            _verify_analysis_batch_workbook(
                payload,
                expected_item_rows=sum(
                    row.roi_id is None
                    for row in rows
                ),
                expected_roi_rows=sum(
                    row.roi_id is not None
                    for row in rows
                ),
                expected_failure_rows=sum(
                    not row.success
                    for row in rows
                ),
            )
        except Exception as exc:
            return AnalysisBatchWorkbookExportResult.failed(
                AnalysisExportFailureCode.VERIFY_FAILED,
                f"批量分析工作簿校验失败：{_error_message(exc)}",
                path=target,
                item_count=item_count,
                success_count=success_count,
                failure_count=failure_count,
            )
        try:
            atomic_write_bytes(target, payload)
        except Exception as exc:
            return AnalysisBatchWorkbookExportResult.failed(
                AnalysisExportFailureCode.ATOMIC_COMMIT_FAILED,
                f"无法写入批量分析工作簿：{_error_message(exc)}",
                path=target,
                item_count=item_count,
                success_count=success_count,
                failure_count=failure_count,
            )
        if not target.is_file() or target.stat().st_size <= 0:
            return AnalysisBatchWorkbookExportResult.failed(
                AnalysisExportFailureCode.VERIFY_FAILED,
                "批量分析工作簿写入后为空。",
                path=target,
                item_count=item_count,
                success_count=success_count,
                failure_count=failure_count,
            )
        return AnalysisBatchWorkbookExportResult.succeeded(
            target,
            item_count=item_count,
            success_count=success_count,
            failure_count=failure_count,
            sheet_count=len(workbook.sheetnames),
        )

    def export_table_csv(
        self,
        artifact: AnalysisArtifact,
        table: AnalysisTable | str | int,
        target_path: str | Path,
    ) -> AnalysisExportResult:
        try:
            target = _normalized_target(target_path, ".csv")
        except (TypeError, ValueError, OSError) as exc:
            return AnalysisExportResult.failed(
                AnalysisExportFailureCode.INVALID_TARGET,
                f"CSV 路径无效：{_error_message(exc)}",
            )
        if not isinstance(artifact, AnalysisArtifact):
            return AnalysisExportResult.failed(
                AnalysisExportFailureCode.ENCODE_FAILED,
                "分析结果对象无效。",
                path=target,
            )
        resolved = _resolve_table(artifact, table)
        if resolved is None:
            return AnalysisExportResult.failed(
                AnalysisExportFailureCode.TABLE_NOT_FOUND,
                "没有找到要导出的分析结果表。",
                path=target,
                artifact_count=1,
            )
        try:
            stream = StringIO(newline="")
            writer = csv.writer(stream, lineterminator="\r\n")
            writer.writerow([_csv_safe_value(column) for column in resolved.columns])
            for row in resolved.rows:
                writer.writerow([_csv_safe_value(value) for value in row])
            payload = stream.getvalue().encode("utf-8-sig")
        except Exception as exc:
            return AnalysisExportResult.failed(
                AnalysisExportFailureCode.ENCODE_FAILED,
                f"无法生成 CSV：{_error_message(exc)}",
                path=target,
                artifact_count=1,
            )
        try:
            atomic_write_bytes(target, payload)
        except Exception as exc:
            return AnalysisExportResult.failed(
                AnalysisExportFailureCode.ATOMIC_COMMIT_FAILED,
                f"无法写入 CSV：{_error_message(exc)}",
                path=target,
                artifact_count=1,
            )
        if not target.is_file() or target.stat().st_size <= 3:
            return AnalysisExportResult.failed(
                AnalysisExportFailureCode.VERIFY_FAILED,
                "CSV 写入后为空。",
                path=target,
                artifact_count=1,
            )
        return AnalysisExportResult.succeeded(
            target,
            artifact_count=1,
            sheet_count=1,
        )

    def export_curve_csv(
        self,
        artifact: AnalysisArtifact,
        curve: AnalysisCurve | str | int,
        target_path: str | Path,
    ) -> AnalysisExportResult:
        try:
            target = _normalized_target(target_path, ".csv")
        except (TypeError, ValueError, OSError) as exc:
            return AnalysisExportResult.failed(
                AnalysisExportFailureCode.INVALID_TARGET,
                f"曲线 CSV 路径无效：{_error_message(exc)}",
            )
        if not isinstance(artifact, AnalysisArtifact):
            return AnalysisExportResult.failed(
                AnalysisExportFailureCode.ENCODE_FAILED,
                "分析结果对象无效。",
                path=target,
            )
        resolved = _resolve_curve(artifact, curve)
        if resolved is None:
            return AnalysisExportResult.failed(
                AnalysisExportFailureCode.CURVE_NOT_FOUND,
                "没有找到要导出的分析曲线。",
                path=target,
                artifact_count=1,
            )
        try:
            stream = StringIO(newline="")
            writer = csv.writer(stream, lineterminator="\r\n")
            x_header = f"X ({resolved.x_unit})" if resolved.x_unit else "X"
            y_header = f"Y ({resolved.y_unit})" if resolved.y_unit else "Y"
            writer.writerow((x_header, y_header))
            for x_value, y_value in zip(resolved.x, resolved.y, strict=True):
                writer.writerow((_csv_safe_value(x_value), _csv_safe_value(y_value)))
            payload = stream.getvalue().encode("utf-8-sig")
        except Exception as exc:
            return AnalysisExportResult.failed(
                AnalysisExportFailureCode.ENCODE_FAILED,
                f"无法生成曲线 CSV：{_error_message(exc)}",
                path=target,
                artifact_count=1,
            )
        try:
            atomic_write_bytes(target, payload)
        except Exception as exc:
            return AnalysisExportResult.failed(
                AnalysisExportFailureCode.ATOMIC_COMMIT_FAILED,
                f"无法写入曲线 CSV：{_error_message(exc)}",
                path=target,
                artifact_count=1,
            )
        return AnalysisExportResult.succeeded(
            target,
            artifact_count=1,
            sheet_count=1,
        )

    def export_portable_package(
        self,
        artifacts: Iterable[AnalysisArtifact],
        target_path: str | Path,
        *,
        asset_root: str | Path | None = None,
        asset_source_paths: Mapping[str, str | Path] | None = None,
        document_names: ArtifactNameMap = None,
        roi_names: ArtifactNameMap = None,
        measurement_names: ArtifactNameMap = None,
        tool_names: ArtifactNameMap = None,
    ) -> AnalysisExportResult:
        frozen = tuple(artifacts)
        try:
            target = _normalized_target(target_path, ".zip")
        except (TypeError, ValueError, OSError) as exc:
            return AnalysisExportResult.failed(
                AnalysisExportFailureCode.INVALID_TARGET,
                f"可搬运分析包路径无效：{_error_message(exc)}",
            )
        if not frozen:
            return AnalysisExportResult.failed(
                AnalysisExportFailureCode.EMPTY_SELECTION,
                "没有可打包的分析结果。",
                path=target,
            )
        if any(not isinstance(item, AnalysisArtifact) for item in frozen):
            return AnalysisExportResult.failed(
                AnalysisExportFailureCode.ENCODE_FAILED,
                "分析结果列表包含不支持的对象。",
                path=target,
            )
        try:
            workbook = _build_workbook(
                frozen,
                document_names=document_names,
                roi_names=roi_names,
                measurement_names=measurement_names,
                tool_names=tool_names,
            )
            workbook_buffer = BytesIO()
            workbook.save(workbook_buffer)
            workbook_payload = workbook_buffer.getvalue()
            _verify_workbook(workbook_payload)
            package_payload = _build_portable_package(
                frozen,
                workbook_payload=workbook_payload,
                asset_root=asset_root,
                asset_source_paths=asset_source_paths,
            )
            _verify_portable_package(package_payload)
        except FileNotFoundError as exc:
            return AnalysisExportResult.failed(
                AnalysisExportFailureCode.ASSET_NOT_FOUND,
                str(exc),
                path=target,
                artifact_count=len(frozen),
            )
        except _AssetHashMismatch as exc:
            return AnalysisExportResult.failed(
                AnalysisExportFailureCode.ASSET_HASH_MISMATCH,
                str(exc),
                path=target,
                artifact_count=len(frozen),
            )
        except Exception as exc:
            return AnalysisExportResult.failed(
                AnalysisExportFailureCode.ENCODE_FAILED,
                f"无法生成可搬运分析包：{_error_message(exc)}",
                path=target,
                artifact_count=len(frozen),
            )
        try:
            atomic_write_bytes(target, package_payload)
        except Exception as exc:
            return AnalysisExportResult.failed(
                AnalysisExportFailureCode.ATOMIC_COMMIT_FAILED,
                f"无法写入可搬运分析包：{_error_message(exc)}",
                path=target,
                artifact_count=len(frozen),
            )
        return AnalysisExportResult.succeeded(
            target,
            artifact_count=len(frozen),
            sheet_count=len(workbook.sheetnames),
        )


def export_analysis_workbook(
    artifacts: Iterable[AnalysisArtifact],
    target_path: str | Path,
    **kwargs: object,
) -> AnalysisExportResult:
    return AnalysisExportService().export_workbook(
        artifacts,
        target_path,
        **kwargs,  # type: ignore[arg-type]
    )


def export_analysis_batch_workbook(
    result: AnalysisBatchResult,
    target_path: str | Path,
    **kwargs: object,
) -> AnalysisBatchWorkbookExportResult:
    return AnalysisExportService().export_batch_workbook(
        result,
        target_path,
        **kwargs,  # type: ignore[arg-type]
    )


def export_analysis_table_csv(
    artifact: AnalysisArtifact,
    table: AnalysisTable | str | int,
    target_path: str | Path,
) -> AnalysisExportResult:
    return AnalysisExportService().export_table_csv(artifact, table, target_path)


def export_analysis_curve_csv(
    artifact: AnalysisArtifact,
    curve: AnalysisCurve | str | int,
    target_path: str | Path,
) -> AnalysisExportResult:
    return AnalysisExportService().export_curve_csv(artifact, curve, target_path)


def export_analysis_portable_package(
    artifacts: Iterable[AnalysisArtifact],
    target_path: str | Path,
    **kwargs: object,
) -> AnalysisExportResult:
    return AnalysisExportService().export_portable_package(
        artifacts,
        target_path,
        **kwargs,  # type: ignore[arg-type]
    )


class _AssetHashMismatch(ValueError):
    pass


def _build_portable_package(
    artifacts: tuple[AnalysisArtifact, ...],
    *,
    workbook_payload: bytes,
    asset_root: str | Path | None,
    asset_source_paths: Mapping[str, str | Path] | None,
) -> bytes:
    files: dict[str, bytes] = {"analysis-results.xlsx": workbook_payload}
    file_records: list[dict[str, object]] = []
    asset_records: list[dict[str, object]] = []
    for artifact_index, artifact in enumerate(artifacts, start=1):
        artifact_segment = (
            f"{artifact_index:04d}-"
            f"{_safe_package_segment(artifact.id, fallback='artifact')}"
        )
        for index, table in enumerate(artifact.tables, start=1):
            name = _safe_package_segment(table.name, fallback="table")
            package_path = f"csv/{artifact_segment}/table-{index:03d}-{name}.csv"
            files[package_path] = _table_csv_bytes(table)
        for index, curve in enumerate(artifact.curves, start=1):
            name = _safe_package_segment(curve.name, fallback="curve")
            package_path = f"csv/{artifact_segment}/curve-{index:03d}-{name}.csv"
            files[package_path] = _curve_csv_bytes(curve)
        for index, reference in enumerate(artifact.assets, start=1):
            source = _resolve_portable_asset(
                reference.path,
                asset_root=asset_root,
                asset_source_paths=asset_source_paths,
            )
            payload = source.read_bytes()
            digest = hashlib.sha256(payload).hexdigest()
            if digest != reference.sha256:
                raise _AssetHashMismatch(
                    f"分析资产 SHA256 不匹配：{reference.path}"
                )
            original_name = PurePosixPath(reference.path).name
            safe_name = _safe_package_segment(
                original_name,
                fallback=f"asset-{index:03d}",
                preserve_suffix=True,
            )
            package_path = (
                f"assets/{artifact_segment}/{index:03d}-{safe_name}"
            )
            if package_path in files:
                raise ValueError(f"分析包路径冲突：{package_path}")
            files[package_path] = payload
            asset_records.append(
                {
                    "artifact_id": artifact.id,
                    "source_path": reference.path,
                    "package_path": package_path,
                    "sha256": digest,
                    "size": len(payload),
                    "media_type": reference.media_type,
                    "schema": reference.metadata.get("schema"),
                }
            )
    for package_path, payload in sorted(files.items()):
        _require_safe_zip_path(package_path)
        file_records.append(
            {
                "path": package_path,
                "sha256": hashlib.sha256(payload).hexdigest(),
                "size": len(payload),
            }
        )
    manifest = {
        "schema_version": 1,
        "created_at": datetime.now(tz=timezone.utc).isoformat(),
        "artifacts": [artifact.to_dict() for artifact in artifacts],
        "assets": asset_records,
        "files": file_records,
    }
    manifest_payload = json.dumps(
        manifest,
        ensure_ascii=False,
        allow_nan=False,
        sort_keys=True,
        separators=(",", ":"),
    ).encode("utf-8")
    buffer = BytesIO()
    with zipfile.ZipFile(
        buffer,
        mode="w",
        compression=zipfile.ZIP_DEFLATED,
        compresslevel=6,
    ) as archive:
        for package_path, payload in sorted(files.items()):
            archive.writestr(package_path, payload)
        archive.writestr("manifest.json", manifest_payload)
    return buffer.getvalue()


def _verify_portable_package(payload: bytes) -> None:
    if not payload:
        raise ValueError("可搬运分析包为空")
    with zipfile.ZipFile(BytesIO(payload), mode="r") as archive:
        names = archive.namelist()
        if len(names) != len(set(names)):
            raise ValueError("可搬运分析包包含重复路径")
        for name in names:
            _require_safe_zip_path(name)
        if "manifest.json" not in names:
            raise ValueError("可搬运分析包缺少 manifest.json")
        manifest = json.loads(archive.read("manifest.json"))
        if (
            not isinstance(manifest, Mapping)
            or manifest.get("schema_version") != 1
            or not isinstance(manifest.get("files"), list)
        ):
            raise ValueError("可搬运分析包 manifest 格式无效")
        for record in manifest["files"]:
            if not isinstance(record, Mapping):
                raise TypeError("manifest.files 必须是对象列表")
            package_path = record.get("path")
            digest = record.get("sha256")
            size = record.get("size")
            if not isinstance(package_path, str) or package_path not in names:
                raise ValueError("manifest 引用了不存在的文件")
            data = archive.read(package_path)
            if len(data) != size:
                raise ValueError(f"分析包文件大小不匹配：{package_path}")
            if hashlib.sha256(data).hexdigest() != digest:
                raise ValueError(f"分析包文件哈希不匹配：{package_path}")


def _resolve_portable_asset(
    asset_path: str,
    *,
    asset_root: str | Path | None,
    asset_source_paths: Mapping[str, str | Path] | None,
) -> Path:
    mapped = dict(asset_source_paths or {}).get(asset_path)
    if mapped is not None:
        candidate = Path(mapped).resolve()
    elif asset_root is not None:
        root = Path(asset_root).resolve()
        candidate = (root / asset_path).resolve()
        try:
            candidate.relative_to(root)
        except ValueError as exc:
            raise ValueError("分析资产路径逃逸项目目录") from exc
    else:
        raise FileNotFoundError(f"缺少分析资产目录：{asset_path}")
    if not candidate.is_file():
        raise FileNotFoundError(f"分析资产不存在：{asset_path}")
    return candidate


def _table_csv_bytes(table: AnalysisTable) -> bytes:
    stream = StringIO(newline="")
    writer = csv.writer(stream, lineterminator="\r\n")
    writer.writerow([_csv_safe_value(column) for column in table.columns])
    for row in table.rows:
        writer.writerow([_csv_safe_value(value) for value in row])
    return stream.getvalue().encode("utf-8-sig")


def _curve_csv_bytes(curve: AnalysisCurve) -> bytes:
    stream = StringIO(newline="")
    writer = csv.writer(stream, lineterminator="\r\n")
    writer.writerow(
        (
            f"X ({curve.x_unit})" if curve.x_unit else "X",
            f"Y ({curve.y_unit})" if curve.y_unit else "Y",
        )
    )
    for x_value, y_value in zip(curve.x, curve.y, strict=True):
        writer.writerow((_csv_safe_value(x_value), _csv_safe_value(y_value)))
    return stream.getvalue().encode("utf-8-sig")


def _safe_package_segment(
    value: object,
    *,
    fallback: str,
    preserve_suffix: bool = False,
) -> str:
    raw = unicodedata.normalize("NFC", str(value))
    suffix = PurePosixPath(raw).suffix if preserve_suffix else ""
    stem = raw[: -len(suffix)] if suffix else raw
    normalized = re.sub(r"[^A-Za-z0-9._-]+", "-", stem).strip(".-")
    normalized = (normalized or fallback)[:96]
    safe_suffix = re.sub(r"[^A-Za-z0-9.]+", "", suffix)[:16]
    return f"{normalized}{safe_suffix}"


def _require_safe_zip_path(value: str) -> None:
    path = PurePosixPath(value)
    if (
        not value
        or path.is_absolute()
        or "\\" in value
        or any(part in {"", ".", ".."} for part in path.parts)
    ):
        raise ValueError(f"分析包路径不安全：{value!r}")


def build_analysis_batch_export_rows(
    result: AnalysisBatchResult,
    *,
    document_names: ArtifactNameMap = None,
    roi_names: ArtifactNameMap = None,
) -> tuple[AnalysisBatchExportRow, ...]:
    """Normalize batch item results into stable workbook rows."""

    if not isinstance(result, AnalysisBatchResult):
        raise TypeError("result 必须是 AnalysisBatchResult")
    rows: list[AnalysisBatchExportRow] = []
    for item in result.item_results:
        if not isinstance(item, AnalysisBatchItemResult):
            raise TypeError("批量分析结果包含无效项目")
        document_id, roi_id = _analysis_batch_export_source_ids(item.item_id)
        executions = tuple(item.executions)
        if item.success and not executions:
            raise ValueError(f"成功项目 {item.item_id} 缺少分析执行结果")
        if not item.success and executions:
            raise ValueError(f"失败项目 {item.item_id} 不应携带分析执行结果")
        document_name = str(
            (document_names or {}).get(
                document_id,
                _analysis_batch_fallback_document_name(item, roi_id),
            )
        )
        roi_name = (
            None
            if roi_id is None
            else str(
                (roi_names or {}).get(
                    roi_id,
                    _analysis_batch_fallback_roi_name(item, roi_id),
                )
            )
        )
        row_executions = executions if executions else (None,)
        for step_index, execution in enumerate(row_executions, start=1):
            rows.append(
                AnalysisBatchExportRow(
                    item_id=item.item_id,
                    display_name=item.display_name,
                    document_id=document_id,
                    document_name=document_name,
                    roi_id=roi_id,
                    roi_name=roi_name,
                    success=item.success,
                    step_index=(step_index if execution is not None else 0),
                    step_count=len(executions),
                    tool_id=(
                        ""
                        if execution is None
                        else execution.tool_spec.tool_id
                    ),
                    tool_name=(
                        ""
                        if execution is None
                        else execution.chinese_name
                    ),
                    algorithm_version=(
                        ""
                        if execution is None
                        else execution.algorithm_version
                    ),
                    scalar_report=(
                        ()
                        if execution is None
                        else tuple(execution.scalar_report)
                    ),
                    array_summaries=(
                        ()
                        if execution is None
                        else tuple(
                            _analysis_batch_array_summary(array)
                            for array in execution.arrays
                        )
                    ),
                    error_type=item.error_type,
                    error_message=item.error_message,
                )
            )
    return tuple(rows)


def _analysis_batch_export_source_ids(
    item_id: str,
) -> tuple[str, str | None]:
    document_id, separator, roi_id = str(item_id).partition("::roi::")
    if not document_id:
        raise ValueError("批量分析项目缺少来源图片 ID")
    if separator and not roi_id:
        raise ValueError("批量分析项目缺少 ROI ID")
    return document_id, (roi_id if separator else None)


def _analysis_batch_fallback_document_name(
    item: AnalysisBatchItemResult,
    roi_id: str | None,
) -> str:
    display_name = str(item.display_name or "").strip()
    if roi_id is not None:
        for separator in (" · ROI：", " · ROI:", " · "):
            if separator in display_name:
                prefix = display_name.split(separator, 1)[0].strip()
                if prefix:
                    return prefix
        return str(item.item_id).split("::roi::", 1)[0]
    return display_name or str(item.item_id)


def _analysis_batch_fallback_roi_name(
    item: AnalysisBatchItemResult,
    roi_id: str,
) -> str:
    display_name = str(item.display_name or "").strip()
    for marker in ("ROI：", "ROI:"):
        if marker in display_name:
            suffix = display_name.rsplit(marker, 1)[1].strip()
            if suffix:
                return suffix
    return roi_id


def _analysis_batch_array_summary(array: object) -> str:
    name = str(getattr(array, "name", "array"))
    shape = tuple(int(value) for value in getattr(array, "shape", ()))
    shape_label = "×".join(str(value) for value in shape) or "标量"
    dtype = str(getattr(array, "dtype", ""))
    byte_count = int(getattr(array, "byte_count", 0))
    return f"{name} [{shape_label}] {dtype} · {byte_count} B"


def _build_analysis_batch_workbook(
    result: AnalysisBatchResult,
    rows: tuple[AnalysisBatchExportRow, ...],
) -> Workbook:
    workbook = Workbook()
    overview = workbook.active
    overview.title = "总览"
    overview.append(
        (
            "批次ID",
            "Generation",
            "配方ID",
            "批次状态",
            "结果项目数",
            "成功数",
            "失败数",
            "整图项目数",
            "ROI项目数",
            "执行结果数",
        )
    )
    overview.append(
        (
            _safe_cell_text(result.request_id),
            result.generation,
            _safe_cell_text(result.recipe_id),
            "已取消" if result.cancelled else "完成",
            len(result.item_results),
            result.success_count,
            result.failure_count,
            sum(
                "::roi::" not in item.item_id
                for item in result.item_results
            ),
            sum(
                "::roi::" in item.item_id
                for item in result.item_results
            ),
            sum(len(item.executions) for item in result.item_results),
        )
    )
    _style_tabular_sheet(overview, frozen_rows=1, auto_filter=True)
    for column in (2, 5, 6, 7, 8, 9, 10):
        overview.cell(row=2, column=column).number_format = "#,##0"
    overview.sheet_view.showGridLines = False

    scalar_names = tuple(
        dict.fromkeys(
            name
            for row in rows
            for name, _value in row.scalar_report
        )
    )
    common_headers = (
        "项目ID",
        "来源图片ID",
        "来源图片",
        "范围类型",
        "ROI ID",
        "ROI",
        "显示名称",
        "状态",
        "配方步骤",
        "步骤总数",
        "分析工具ID",
        "分析工具",
        "算法版本",
        "数组数量",
        "数组摘要",
    )
    result_headers = common_headers + tuple(
        f"指标·{name}"
        for name in scalar_names
    )

    by_image = workbook.create_sheet("逐图片")
    by_image.append(result_headers)
    for row in rows:
        if row.roi_id is not None:
            continue
        by_image.append(_analysis_batch_export_values(row, scalar_names))
    _style_tabular_sheet(by_image, frozen_rows=1, auto_filter=True)
    _style_analysis_batch_data_rows(by_image, metric_start=16)

    by_roi = workbook.create_sheet("逐 ROI")
    by_roi.append(result_headers)
    for row in rows:
        if row.roi_id is not None:
            by_roi.append(_analysis_batch_export_values(row, scalar_names))
    _style_tabular_sheet(by_roi, frozen_rows=1, auto_filter=True)
    _style_analysis_batch_data_rows(by_roi, metric_start=16)

    failures = workbook.create_sheet("失败明细")
    failures.append(
        (
            "项目ID",
            "来源图片ID",
            "来源图片",
            "范围类型",
            "ROI ID",
            "ROI",
            "显示名称",
            "错误类型",
            "错误消息",
        )
    )
    for row in rows:
        if row.success:
            continue
        failures.append(
            (
                _safe_cell_text(row.item_id),
                _safe_cell_text(row.document_id),
                _safe_cell_text(row.document_name),
                row.scope,
                _safe_cell_text(row.roi_id or ""),
                _safe_cell_text(row.roi_name or ""),
                _safe_cell_text(row.display_name),
                _safe_cell_text(row.error_type or ""),
                _safe_cell_text(row.error_message or ""),
            )
        )
    _style_tabular_sheet(failures, frozen_rows=1, auto_filter=True)
    _style_analysis_batch_data_rows(failures, metric_start=None)

    for sheet in (by_image, by_roi, failures):
        sheet.sheet_view.showGridLines = False
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    return workbook


def _analysis_batch_export_values(
    row: AnalysisBatchExportRow,
    scalar_names: tuple[str, ...],
) -> tuple[object, ...]:
    scalars = dict(row.scalar_report)
    return (
        _safe_cell_text(row.item_id),
        _safe_cell_text(row.document_id),
        _safe_cell_text(row.document_name),
        row.scope,
        _safe_cell_text(row.roi_id or ""),
        _safe_cell_text(row.roi_name or ""),
        _safe_cell_text(row.display_name),
        "成功" if row.success else "失败",
        row.step_label,
        row.step_count,
        _safe_cell_text(row.tool_id),
        _safe_cell_text(row.tool_name),
        _safe_cell_text(row.algorithm_version),
        len(row.array_summaries),
        _safe_cell_text("；".join(row.array_summaries)),
        *(
            _safe_cell_value(scalars.get(name))
            for name in scalar_names
        ),
    )


def _style_analysis_batch_data_rows(
    sheet,
    *,
    metric_start: int | None,
) -> None:
    success_fill = PatternFill(fill_type="solid", fgColor="E2F0D9")
    failure_fill = PatternFill(fill_type="solid", fgColor="FCE4D6")
    status_column = 8 if sheet.title != "失败明细" else None
    for row_index in range(2, sheet.max_row + 1):
        if status_column is not None:
            status = sheet.cell(row=row_index, column=status_column)
            status.fill = (
                success_fill
                if status.value == "成功"
                else failure_fill
            )
        elif sheet.title == "失败明细":
            sheet.cell(row=row_index, column=8).fill = failure_fill
            sheet.cell(row=row_index, column=9).fill = failure_fill
        if metric_start is None:
            continue
        sheet.cell(row=row_index, column=14).number_format = "#,##0"
        for column_index in range(metric_start, sheet.max_column + 1):
            cell = sheet.cell(row=row_index, column=column_index)
            if isinstance(cell.value, bool):
                continue
            if isinstance(cell.value, int):
                cell.number_format = "#,##0"
            elif isinstance(cell.value, float):
                cell.number_format = "0.###############"


def _verify_analysis_batch_workbook(
    payload: bytes,
    *,
    expected_item_rows: int,
    expected_roi_rows: int,
    expected_failure_rows: int,
) -> None:
    if not payload:
        raise ValueError("批量分析工作簿内容为空")
    workbook = load_workbook(BytesIO(payload), read_only=True, data_only=False)
    try:
        expected_sheets = ["总览", "逐图片", "逐 ROI", "失败明细"]
        if workbook.sheetnames != expected_sheets:
            raise ValueError("批量分析工作簿页签不完整")
        if workbook["总览"].max_row != 2:
            raise ValueError("批量分析总览页格式无效")
        if workbook["逐图片"].max_row != expected_item_rows + 1:
            raise ValueError("逐图片结果行数量不匹配")
        if workbook["逐 ROI"].max_row != expected_roi_rows + 1:
            raise ValueError("逐 ROI 结果行数量不匹配")
        if workbook["失败明细"].max_row != expected_failure_rows + 1:
            raise ValueError("失败明细结果行数量不匹配")
        if workbook["总览"]["A1"].value != "批次ID":
            raise ValueError("批量分析总览页表头无效")
        if workbook["失败明细"]["A1"].value != "项目ID":
            raise ValueError("批量分析失败明细页表头无效")
    finally:
        workbook.close()


def _build_workbook(
    artifacts: tuple[AnalysisArtifact, ...],
    *,
    document_names: ArtifactNameMap,
    roi_names: ArtifactNameMap,
    measurement_names: ArtifactNameMap,
    tool_names: ArtifactNameMap,
) -> Workbook:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "分析摘要"
    summary_headers = (
        "分析结果ID",
        "状态",
        "失效原因",
        "来源文档",
        "来源像素修订",
        "来源对象",
        "分析工具",
        "算法版本",
        "标量摘要",
        "生成时间",
        "标定签名",
    )
    summary.append(summary_headers)
    for artifact in artifacts:
        summary.append(
            (
                _safe_cell_text(artifact.id),
                "当前" if artifact.status is AnalysisArtifactStatus.CURRENT else "已失效",
                _safe_cell_text(artifact.stale_reason or ""),
                _safe_cell_text(
                    (document_names or {}).get(
                        artifact.source_document_id,
                        artifact.source_document_id,
                    )
                ),
                artifact.source_pixel_revision,
                _safe_cell_text(
                    _source_reference_label(
                        artifact,
                        roi_names=roi_names,
                        measurement_names=measurement_names,
                    )
                ),
                _safe_cell_text(_tool_label(artifact, tool_names)),
                _safe_cell_text(artifact.tool_version),
                _safe_cell_text(_json_text(artifact.scalars)),
                _safe_cell_text(artifact.created_at),
                _safe_cell_text(artifact.calibration_signature or "未标定"),
            )
        )
    _style_tabular_sheet(summary, frozen_rows=1, auto_filter=True)

    parameters = workbook.create_sheet("参数与来源")
    parameters.append(("分析结果ID", "分析工具", "信息类型", "名称", "值"))
    for artifact in artifacts:
        tool_label = _tool_label(artifact, tool_names)
        source_rows = (
            ("来源", "来源文档ID", artifact.source_document_id),
            ("来源", "来源像素修订", artifact.source_pixel_revision),
            (
                "来源",
                "来源对象",
                _source_reference_label(
                    artifact,
                    roi_names=roi_names,
                    measurement_names=measurement_names,
                ),
            ),
            ("来源", "标定签名", artifact.calibration_signature or "未标定"),
            ("来源", "生成时间", artifact.created_at),
        )
        for information_type, name, value in source_rows:
            parameters.append(
                (
                    _safe_cell_text(artifact.id),
                    _safe_cell_text(tool_label),
                    information_type,
                    name,
                    _safe_cell_value(value),
                )
            )
        for name, value in artifact.parameters.items():
            display_name = (
                "输出字段选择"
                if name == ANALYSIS_OUTPUT_FIELDS_PARAMETER
                else name
            )
            display_value = value
            if (
                name == ANALYSIS_OUTPUT_FIELDS_PARAMETER
                and isinstance(value, list)
            ):
                schema = analysis_output_field_schema(artifact.tool_id)
                if schema is not None:
                    field_names = {
                        field.key: field.chinese_name for field in schema.fields
                    }
                    display_value = "、".join(
                        field_names.get(str(item), str(item)) for item in value
                    ) or "仅必要审计字段"
            parameters.append(
                (
                    _safe_cell_text(artifact.id),
                    _safe_cell_text(tool_label),
                    "参数",
                    _safe_cell_text(display_name),
                    _safe_cell_value(display_value),
                )
            )
        for warning in artifact.warnings:
            parameters.append(
                (
                    _safe_cell_text(artifact.id),
                    _safe_cell_text(tool_label),
                    "提示",
                    "分析提示",
                    _safe_cell_text(warning),
                )
            )
    _style_tabular_sheet(parameters, frozen_rows=1, auto_filter=True)

    grouped: dict[str, list[AnalysisArtifact]] = {}
    for artifact in artifacts:
        grouped.setdefault(artifact.tool_id, []).append(artifact)
    used_names = set(workbook.sheetnames)
    for tool_id, tool_artifacts in grouped.items():
        label = _tool_label(tool_artifacts[0], tool_names)
        title = _unique_sheet_title(f"详情_{label}", used_names)
        used_names.add(title)
        sheet = workbook.create_sheet(title)
        sheet.append(
            (
                "分析结果ID",
                "状态",
                "结果类型",
                "结果名称",
                "行号",
                "字段",
                "值",
                "X",
                "Y",
                "X单位",
                "Y单位",
            )
        )
        for artifact in tool_artifacts:
            status = (
                "当前"
                if artifact.status is AnalysisArtifactStatus.CURRENT
                else "已失效"
            )
            for name, value in artifact.scalars.items():
                sheet.append(
                    (
                        _safe_cell_text(artifact.id),
                        status,
                        "标量",
                        "分析摘要",
                        None,
                        _safe_cell_text(name),
                        _safe_cell_value(value),
                        None,
                        None,
                        "",
                        "",
                    )
                )
            for table in artifact.tables:
                for row_number, row in enumerate(table.rows, start=1):
                    for column, value in zip(table.columns, row, strict=True):
                        sheet.append(
                            (
                                _safe_cell_text(artifact.id),
                                status,
                                "表格",
                                _safe_cell_text(table.name),
                                row_number,
                                _safe_cell_text(column),
                                _safe_cell_value(value),
                                None,
                                None,
                                "",
                                "",
                            )
                        )
            for curve in artifact.curves:
                _append_curve_rows(sheet, artifact, status, curve)
            for asset in artifact.assets:
                sheet.append(
                    (
                        _safe_cell_text(artifact.id),
                        status,
                        "资产",
                        _safe_cell_text(asset.kind.value),
                        None,
                        "路径",
                        _safe_cell_text(asset.path),
                        None,
                        None,
                        "",
                        "",
                    )
                )
        _style_tabular_sheet(sheet, frozen_rows=1, auto_filter=True)
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    return workbook


def _append_curve_rows(
    sheet,
    artifact: AnalysisArtifact,
    status: str,
    curve: AnalysisCurve,
) -> None:
    for index, (x_value, y_value) in enumerate(zip(curve.x, curve.y, strict=True), start=1):
        sheet.append(
            (
                _safe_cell_text(artifact.id),
                status,
                "曲线",
                _safe_cell_text(curve.name),
                index,
                "",
                None,
                x_value,
                y_value,
                _safe_cell_text(curve.x_unit),
                _safe_cell_text(curve.y_unit),
            )
        )


def _style_tabular_sheet(sheet, *, frozen_rows: int, auto_filter: bool) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="2A9D8F")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    sheet.freeze_panes = f"A{frozen_rows + 1}"
    if auto_filter and sheet.max_row >= 1:
        sheet.auto_filter.ref = sheet.dimensions
    for column_index in range(1, sheet.max_column + 1):
        maximum = 0
        for row_index in range(1, min(sheet.max_row, 300) + 1):
            value = sheet.cell(row=row_index, column=column_index).value
            maximum = max(maximum, len(str(value or "")))
        sheet.column_dimensions[get_column_letter(column_index)].width = min(
            48,
            max(10, maximum + 2),
        )
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _verify_workbook(payload: bytes) -> None:
    if not payload:
        raise ValueError("工作簿内容为空")
    workbook = load_workbook(BytesIO(payload), read_only=True, data_only=False)
    try:
        if workbook.sheetnames[:2] != ["分析摘要", "参数与来源"]:
            raise ValueError("缺少分析摘要或参数与来源页")
        if len(workbook.sheetnames) < 3:
            raise ValueError("缺少工具详情页")
        if workbook["分析摘要"].max_row < 2:
            raise ValueError("分析摘要没有数据")
    finally:
        workbook.close()


def _resolve_table(
    artifact: AnalysisArtifact,
    table: AnalysisTable | str | int,
) -> AnalysisTable | None:
    if isinstance(table, AnalysisTable):
        return table if table in artifact.tables else None
    if isinstance(table, int) and not isinstance(table, bool):
        return artifact.tables[table] if 0 <= table < len(artifact.tables) else None
    normalized_name = str(table)
    return next((item for item in artifact.tables if item.name == normalized_name), None)


def _resolve_curve(
    artifact: AnalysisArtifact,
    curve: AnalysisCurve | str | int,
) -> AnalysisCurve | None:
    if isinstance(curve, AnalysisCurve):
        return curve if curve in artifact.curves else None
    if isinstance(curve, int) and not isinstance(curve, bool):
        return artifact.curves[curve] if 0 <= curve < len(artifact.curves) else None
    normalized_name = str(curve)
    return next((item for item in artifact.curves if item.name == normalized_name), None)


def _source_reference_label(
    artifact: AnalysisArtifact,
    *,
    roi_names: ArtifactNameMap,
    measurement_names: ArtifactNameMap,
) -> str:
    reference = artifact.source_reference
    if reference is None:
        return "整张图片"
    if reference.kind is AnalysisObjectKind.ROI:
        label = (roi_names or {}).get(reference.object_id, reference.object_id)
        return f"ROI：{label}（修订 {reference.revision}）"
    label = (measurement_names or {}).get(reference.object_id, reference.object_id)
    return f"测量对象：{label}（修订 {reference.revision}）"


def _tool_label(
    artifact: AnalysisArtifact,
    tool_names: ArtifactNameMap,
) -> str:
    explicit = (tool_names or {}).get(artifact.tool_id)
    if explicit:
        return explicit
    parameter_name = artifact.parameters.get("tool_name")
    if isinstance(parameter_name, str) and parameter_name.strip():
        return parameter_name.strip()
    return _TOOL_NAMES.get(artifact.tool_id, artifact.tool_id)


def _normalized_target(value: str | Path, suffix: str) -> Path:
    target = Path(value)
    if not target.name or target.name in {".", ".."}:
        raise ValueError("导出路径无效")
    return target.with_suffix(suffix)


def _unique_sheet_title(raw_title: str, used_names: set[str]) -> str:
    normalized = unicodedata.normalize("NFC", str(raw_title))
    normalized = _INVALID_SHEET_CHARS.sub("_", normalized).strip().strip("'")
    normalized = normalized or "详情"
    base = normalized[:31]
    used_folded = {unicodedata.normalize("NFC", name).casefold() for name in used_names}
    if base.casefold() not in used_folded:
        return base
    counter = 2
    while True:
        suffix = f"_{counter}"
        candidate = f"{base[: 31 - len(suffix)]}{suffix}"
        if candidate.casefold() not in used_folded:
            return candidate
        counter += 1


def _safe_cell_value(value: object) -> object:
    if value is None or isinstance(value, (bool, int, float)):
        return value
    if isinstance(value, str):
        return _safe_cell_text(value)
    return _safe_cell_text(_json_text(value))


def _safe_cell_text(value: object) -> str:
    text = str(value)
    return f"'{text}" if text.startswith(_FORMULA_PREFIXES) else text


def _csv_safe_value(value: JsonScalar) -> JsonScalar:
    if isinstance(value, str) and value.startswith(_FORMULA_PREFIXES):
        return f"'{value}"
    return value


def _json_text(value: object) -> str:
    return json.dumps(
        value,
        ensure_ascii=False,
        allow_nan=False,
        sort_keys=True,
        separators=(",", ":"),
    )


def _error_message(error: Exception) -> str:
    return str(error).strip() or type(error).__name__


__all__ = [
    "AnalysisBatchExportRow",
    "AnalysisBatchWorkbookExportResult",
    "AnalysisExportFailureCode",
    "AnalysisExportResult",
    "AnalysisExportService",
    "build_analysis_batch_export_rows",
    "export_analysis_batch_workbook",
    "export_analysis_curve_csv",
    "export_analysis_portable_package",
    "export_analysis_table_csv",
    "export_analysis_workbook",
]
