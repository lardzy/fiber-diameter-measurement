from __future__ import annotations

from pathlib import Path
import shutil
import sys
from typing import Any

from fdm.content_experiment import (
    ContentExperimentSession,
    ContentRecordKind,
    content_session_stats,
)
from fdm.models import project_assets_root
from fdm.runtime_logging import append_runtime_log
from fdm.settings import runtime_directory


RAW_SHEET_NAME = "原始数据"
REPORT_SHEET_NAME = "特种毛(报出)"
CONTENT_WORKBOOK_RELATIVE_DIR = "content_experiments"


def content_template_path() -> Path:
    return runtime_directory() / "content-templates" / "sheet.xlt"


class ContentWorkbookService:
    """Keeps the content-experiment workbook synchronized.

    On Windows with Excel installed this service opens a visible workbook via COM
    so the user can see updates live. Everywhere else it keeps an .xlsx snapshot
    with openpyxl; that snapshot is the source for project resume/export.
    """

    def __init__(self) -> None:
        self._excel_app: Any | None = None
        self._com_workbook: Any | None = None
        self._pythoncom: Any | None = None
        self._py_workbook: Any | None = None
        self._mode = ""
        self._last_warning = ""
        self._last_fiber_ids: list[str] = []
        self._last_basic_values: tuple[str, str, str] | None = None
        self._last_diameter_ids_by_fiber: dict[str, list[str]] = {}

    @property
    def mode(self) -> str:
        return self._mode

    @property
    def last_warning(self) -> str:
        return self._last_warning

    def is_open(self) -> bool:
        return self._com_workbook is not None or self._py_workbook is not None

    def open_session(self, session: ContentExperimentSession, *, project_path: str | Path | None = None) -> str:
        self.close()
        self._last_warning = ""
        snapshot = self.snapshot_path(session, project_path) if project_path is not None else None
        if sys.platform.startswith("win"):
            try:
                self._open_com(snapshot=snapshot)
                self._mode = "excel"
                self._last_warning = ""
                session.workbook_mode = self._mode
                self.sync_session(session)
                return self._mode
            except Exception as exc:
                self._last_warning = f"Excel 模板打开失败，已切换为 xlsx 快照模式：{exc}"
                append_runtime_log("Content workbook COM open failed", repr(exc))
                self.close()
        self._open_openpyxl(snapshot=snapshot)
        self._mode = "xlsx"
        session.workbook_mode = self._mode
        self.sync_session(session)
        return self._mode

    def close(self) -> None:
        if self._com_workbook is not None:
            try:
                self._com_workbook.Close(SaveChanges=False)
            except Exception:
                pass
        self._com_workbook = None
        self._excel_app = None
        if self._pythoncom is not None:
            try:
                self._pythoncom.CoUninitialize()
            except Exception:
                pass
        self._pythoncom = None
        self._py_workbook = None
        self._mode = ""
        self._last_fiber_ids = []
        self._last_basic_values = None
        self._last_diameter_ids_by_fiber = {}

    def snapshot_path(self, session: ContentExperimentSession, project_path: str | Path | None) -> Path | None:
        if project_path is None:
            return None
        assets_root = project_assets_root(project_path)
        if session.workbook_snapshot_relpath:
            return (assets_root / session.workbook_snapshot_relpath).resolve()
        return (assets_root / CONTENT_WORKBOOK_RELATIVE_DIR / f"{session.id}.xlsx").resolve()

    def save_snapshot(self, session: ContentExperimentSession, project_path: str | Path) -> Path:
        snapshot = project_assets_root(project_path) / CONTENT_WORKBOOK_RELATIVE_DIR / f"{session.id}.xlsx"
        snapshot.parent.mkdir(parents=True, exist_ok=True)
        self.sync_session(session)
        if self._com_workbook is not None:
            self._com_workbook.SaveCopyAs(str(snapshot))
        elif self._py_workbook is not None:
            self._py_workbook.save(str(snapshot))
        else:
            self._open_openpyxl(snapshot=None)
            self.sync_session(session)
            self._py_workbook.save(str(snapshot))
        session.workbook_snapshot_relpath = snapshot.relative_to(project_assets_root(project_path)).as_posix()
        session.workbook_mode = self._mode or "xlsx"
        return snapshot

    def save_as(self, session: ContentExperimentSession, output_path: str | Path) -> Path:
        target = Path(output_path)
        target.parent.mkdir(parents=True, exist_ok=True)
        self.sync_session(session)
        if self._com_workbook is not None:
            self._com_workbook.SaveCopyAs(str(target))
        elif self._py_workbook is not None:
            self._py_workbook.save(str(target))
        else:
            self._open_openpyxl(snapshot=None)
            self.sync_session(session)
            self._py_workbook.save(str(target))
        return target

    def sync_session(self, session: ContentExperimentSession) -> None:
        if self._com_workbook is not None:
            self._sync_com(session)
            return
        if self._py_workbook is None:
            self._open_openpyxl(snapshot=None)
        self._sync_openpyxl(session)

    def _open_com(self, *, snapshot: Path | None) -> None:
        import win32com.client  # type: ignore[import-not-found]

        try:
            import pythoncom  # type: ignore[import-not-found]

            pythoncom.CoInitialize()
            self._pythoncom = pythoncom
        except Exception:
            self._pythoncom = None

        excel_errors: list[str] = []
        excel = None
        prog_ids = ("Excel.Application", "Excel.Application.16", "Excel.Application.15", "Excel.Application.14")
        for factory_name in ("DispatchEx", "Dispatch", "EnsureDispatch"):
            factory = getattr(win32com.client, factory_name, None)
            if not callable(factory):
                continue
            for prog_id in prog_ids:
                try:
                    excel = factory(prog_id)
                    break
                except Exception as exc:  # noqa: BLE001
                    excel_errors.append(f"{factory_name}({prog_id}): {exc}")
            if excel is not None:
                break
        if excel is None:
            detail = "; ".join(excel_errors) or "win32com.client 未提供 Dispatch/DispatchEx。"
            raise RuntimeError(
                "无法创建 Excel.Application COM 对象。"
                "请确认当前 Python 环境已安装 pywin32，且桌面版 Microsoft Excel 已完成 COM 注册。"
                f"诊断: {detail}"
            )
        excel.Visible = True
        if snapshot is not None and snapshot.exists():
            workbook = excel.Workbooks.Open(str(snapshot))
        else:
            template = content_template_path()
            if not template.exists():
                raise FileNotFoundError(f"含量试验模板不存在: {template}")
            workbook_errors: list[str] = []
            workbook = None
            try:
                workbook = excel.Workbooks.Add(Template=str(template))
            except Exception as exc:  # noqa: BLE001
                workbook_errors.append(f"Add(Template=...): {exc}")
            if workbook is None:
                try:
                    workbook = excel.Workbooks.Add(str(template))
                except Exception as exc:  # noqa: BLE001
                    workbook_errors.append(f"Add(path): {exc}")
            if workbook is None:
                try:
                    workbook = excel.Workbooks.Open(str(template))
                except Exception as exc:  # noqa: BLE001
                    workbook_errors.append(f"Open(template): {exc}")
            if workbook is None:
                raise RuntimeError("无法用 Excel 打开含量试验模板。" + "；".join(workbook_errors))
        self._excel_app = excel
        self._com_workbook = workbook

    def _open_openpyxl(self, *, snapshot: Path | None) -> None:
        from openpyxl import Workbook, load_workbook

        if snapshot is not None and snapshot.exists():
            try:
                self._py_workbook = load_workbook(snapshot)
                return
            except Exception:
                pass
        self._py_workbook = Workbook()
        default = self._py_workbook.active
        default.title = RAW_SHEET_NAME
        if REPORT_SHEET_NAME not in self._py_workbook.sheetnames:
            self._py_workbook.create_sheet(REPORT_SHEET_NAME)
        self._ensure_fallback_labels()

    def _sync_openpyxl(self, session: ContentExperimentSession) -> None:
        workbook = self._py_workbook
        if workbook is None:
            return
        raw_sheet = workbook[RAW_SHEET_NAME] if RAW_SHEET_NAME in workbook.sheetnames else workbook.create_sheet(RAW_SHEET_NAME)
        report_sheet = workbook[REPORT_SHEET_NAME] if REPORT_SHEET_NAME in workbook.sheetnames else workbook.create_sheet(REPORT_SHEET_NAME)
        report_sheet["D2"] = session.sample_name
        report_sheet["D3"] = session.sample_id
        report_sheet["D7"] = session.operator
        if self._can_incremental_sync(session):
            self._write_incremental_values(raw_sheet, session, kind="openpyxl")
        else:
            self._write_sheet_values(raw_sheet, session, kind="openpyxl")
        self._remember_sync_state(session)

    def _sync_com(self, session: ContentExperimentSession) -> None:
        workbook = self._com_workbook
        if workbook is None:
            return
        raw_sheet = workbook.Worksheets(RAW_SHEET_NAME)
        report_sheet = workbook.Worksheets(REPORT_SHEET_NAME)
        report_sheet.Range("D2").Value = session.sample_name
        report_sheet.Range("D3").Value = session.sample_id
        report_sheet.Range("D7").Value = session.operator
        if self._can_incremental_sync(session):
            self._write_incremental_values(raw_sheet, session, kind="com")
        else:
            self._write_sheet_values(raw_sheet, session, kind="com")
        self._remember_sync_state(session)

    def _write_sheet_values(self, sheet, session: ContentExperimentSession, *, kind: str) -> None:
        stats = content_session_stats(session)
        counts = {item.fiber.id: item.count for item in stats}
        measured = {item.fiber.id: item.measured for item in stats}
        averages = {item.fiber.id: item.average_diameter for item in stats}
        mean_sq = {item.fiber.id: item.mean_diameter_squared for item in stats}
        diameter_values: dict[str, list[float]] = {fiber.id: [] for fiber in session.fibers}
        for record in session.records:
            if record.kind != ContentRecordKind.DIAMETER:
                continue
            value = record.diameter_unit if record.diameter_unit is not None else record.diameter_px
            if value is not None and record.fiber_id in diameter_values:
                diameter_values[record.fiber_id].append(float(value))

        for index in range(8):
            column = 4 + index
            fiber = session.fibers[index] if index < len(session.fibers) else None
            self._set_cell(sheet, 6, column, fiber.name if fiber is not None else "", kind=kind)
            self._set_cell(sheet, 7, column, fiber.density if fiber is not None and fiber.density is not None else "", kind=kind)
            self._set_cell(sheet, 8, column, counts.get(fiber.id, 0) if fiber is not None else "", kind=kind)
            self._set_cell(sheet, 9, column, measured.get(fiber.id, 0) if fiber is not None else "", kind=kind)
            self._set_cell(sheet, 10, column, averages.get(fiber.id) if fiber is not None and averages.get(fiber.id) is not None else "", kind=kind)
            self._set_cell(sheet, 14, column, mean_sq.get(fiber.id) if fiber is not None and mean_sq.get(fiber.id) is not None else "", kind=kind)
            self._clear_diameter_column(sheet, column, kind=kind)
            if fiber is None:
                continue
            for row_offset, value in enumerate(diameter_values.get(fiber.id, []), start=25):
                self._set_cell(sheet, row_offset, column, value, kind=kind)

    def _write_incremental_values(self, sheet, session: ContentExperimentSession, *, kind: str) -> None:
        stats = content_session_stats(session)
        stats_by_id = {item.fiber.id: item for item in stats}
        diameter_records = self._diameter_records_by_fiber(session)
        for index, fiber in enumerate(session.fibers):
            column = 4 + index
            stats_item = stats_by_id[fiber.id]
            self._set_cell(sheet, 8, column, stats_item.count, kind=kind)
            self._set_cell(sheet, 9, column, stats_item.measured, kind=kind)
            self._set_cell(sheet, 10, column, stats_item.average_diameter if stats_item.average_diameter is not None else "", kind=kind)
            self._set_cell(sheet, 14, column, stats_item.mean_diameter_squared if stats_item.mean_diameter_squared is not None else "", kind=kind)

            previous_ids = self._last_diameter_ids_by_fiber.get(fiber.id, [])
            current_records = diameter_records.get(fiber.id, [])
            current_ids = [record.id for record in current_records]
            if current_ids == previous_ids:
                continue
            if len(current_ids) == len(previous_ids) + 1 and current_ids[: len(previous_ids)] == previous_ids:
                record = current_records[-1]
                value = record.diameter_unit if record.diameter_unit is not None else record.diameter_px
                self._set_cell(sheet, 24 + len(current_ids), column, value if value is not None else "", kind=kind)
                continue
            self._clear_diameter_column(sheet, column, kind=kind)
            for row_offset, record in enumerate(current_records, start=25):
                value = record.diameter_unit if record.diameter_unit is not None else record.diameter_px
                self._set_cell(sheet, row_offset, column, value if value is not None else "", kind=kind)

    def _ensure_fallback_labels(self) -> None:
        if self._py_workbook is None:
            return
        raw = self._py_workbook[RAW_SHEET_NAME]
        report = self._py_workbook[REPORT_SHEET_NAME]
        report["C2"] = "样品名称"
        report["C3"] = "样品编号"
        report["C7"] = "试验者"
        raw["C6"] = "纤维类别"
        raw["C7"] = "比重"
        raw["C8"] = "记数根数"
        raw["C9"] = "实测根数"
        raw["C10"] = "平均直径"
        raw["C14"] = "直径平方均值"
        raw["C25"] = "直径明细"

    def _set_cell(self, sheet, row: int, column: int, value: object, *, kind: str) -> None:
        if kind == "com":
            sheet.Cells(row, column).Value = value
        else:
            sheet.cell(row=row, column=column).value = value

    def _clear_diameter_column(self, sheet, column: int, *, kind: str) -> None:
        if kind == "com":
            sheet.Range(sheet.Cells(25, column), sheet.Cells(2024, column)).ClearContents()
            return
        max_row = max(2024, int(getattr(sheet, "max_row", 2024) or 2024))
        for row in range(25, max_row + 1):
            sheet.cell(row=row, column=column).value = None

    def _can_incremental_sync(self, session: ContentExperimentSession) -> bool:
        if not self._last_fiber_ids:
            return False
        return self._last_fiber_ids == [fiber.id for fiber in session.fibers]

    def _remember_sync_state(self, session: ContentExperimentSession) -> None:
        self._last_fiber_ids = [fiber.id for fiber in session.fibers]
        self._last_basic_values = (session.operator, session.sample_id, session.sample_name)
        self._last_diameter_ids_by_fiber = {
            fiber_id: [record.id for record in records]
            for fiber_id, records in self._diameter_records_by_fiber(session).items()
        }

    @staticmethod
    def _diameter_records_by_fiber(session: ContentExperimentSession) -> dict[str, list[Any]]:
        records: dict[str, list[Any]] = {fiber.id: [] for fiber in session.fibers}
        for record in session.records:
            if record.kind == ContentRecordKind.DIAMETER and record.fiber_id in records:
                records[record.fiber_id].append(record)
        return records


def copy_template_to_runtime(source: str | Path, target: str | Path | None = None) -> Path:
    output_path = Path(target) if target is not None else content_template_path()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copyfile(source, output_path)
    return output_path
