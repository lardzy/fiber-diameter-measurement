from __future__ import annotations

import hashlib
import json
import os
from pathlib import Path, PurePosixPath
import subprocess
import sys
import tempfile
from typing import Any

from fdm.area_worker_protocol import AREA_WORKER_PROTOCOL, AREA_WORKER_PROTOCOL_VERSION


RELEASE_MANIFEST_FILENAME = "release-manifest.json"
BUILD_ID_FILENAME = "build-id.txt"
DEVELOPMENT_FEATURES = frozenset(
    {"measurement", "capture", "digital-slide", "area-inference", "magic-segmentation"}
)


def sha256_file(path: str | Path) -> str:
    digest = hashlib.sha256()
    with Path(path).open("rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def release_root() -> Path:
    override = os.environ.get("FDM_SELF_CHECK_ROOT", "").strip()
    if override:
        return Path(override).expanduser().resolve()
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parents[2]


def _safe_manifest_path(root: Path, token: object) -> Path | None:
    normalized = str(token or "").replace("\\", "/").strip()
    pure_path = PurePosixPath(normalized)
    if not normalized or pure_path.is_absolute() or ".." in pure_path.parts:
        return None
    candidate = root.joinpath(*pure_path.parts)
    try:
        candidate.resolve().relative_to(root.resolve())
    except (OSError, ValueError):
        return None
    return candidate


def verify_release_manifest(
    app_root: str | Path,
    *,
    expected_profile: str | None = None,
    expected_version: str | None = None,
    expected_commit: str | None = None,
    expected_excluded_components: tuple[str, ...] | list[str] | None = None,
    expected_included_components: tuple[str, ...] | list[str] | None = None,
    require_clean_source: bool = False,
    require_clean_build: bool = False,
    reject_extra_files: bool = False,
) -> dict[str, Any]:
    root = Path(app_root)
    errors: list[str] = []
    warnings: list[str] = []
    manifest_path = root / RELEASE_MANIFEST_FILENAME
    result: dict[str, Any] = {
        "ok": False,
        "root": str(root),
        "manifest": str(manifest_path),
        "profile": "",
        "version": "",
        "build_id": "",
        "source_commit": "",
        "features": [],
        "excluded_components": [],
        "included_components": [],
        "dependency_versions": {},
        "checked_files": 0,
        "errors": errors,
        "warnings": warnings,
    }
    try:
        payload = json.loads(manifest_path.read_text(encoding="utf-8"))
    except FileNotFoundError:
        errors.append(f"missing {RELEASE_MANIFEST_FILENAME}")
        return result
    except (OSError, json.JSONDecodeError) as exc:
        errors.append(f"invalid {RELEASE_MANIFEST_FILENAME}: {exc}")
        return result
    if not isinstance(payload, dict):
        errors.append("release manifest root must be a JSON object")
        return result

    schema_version = payload.get("schema_version")
    if schema_version != 1:
        errors.append(f"unsupported release manifest schema: {schema_version!r}")
    profile = str(payload.get("profile", ""))
    version = str(payload.get("version", ""))
    build_id = str(payload.get("build_id", ""))
    source_commit = str(payload.get("source_commit", ""))
    features_payload = payload.get("features", [])
    if not isinstance(features_payload, list) or any(not isinstance(item, str) for item in features_payload):
        errors.append("release manifest features must be a list of strings")
        features_payload = []
    features = [str(item).strip() for item in features_payload if str(item).strip()]
    excluded_payload = payload.get("excluded_components", [])
    included_payload = payload.get("included_components", [])
    if not isinstance(excluded_payload, list) or any(
        not isinstance(item, str) for item in excluded_payload
    ):
        errors.append("release manifest excluded_components must be a list of strings")
        excluded_payload = []
    if not isinstance(included_payload, list) or any(
        not isinstance(item, str) for item in included_payload
    ):
        errors.append("release manifest included_components must be a list of strings")
        included_payload = []
    excluded_components = [str(item).strip() for item in excluded_payload if str(item).strip()]
    included_components = [str(item).strip() for item in included_payload if str(item).strip()]
    dependency_versions = payload.get("dependency_versions", {})
    if not isinstance(dependency_versions, dict) or any(
        not isinstance(key, str) or not isinstance(value, str)
        for key, value in dependency_versions.items()
    ):
        errors.append("release manifest dependency_versions must be a string mapping")
        dependency_versions = {}
    result.update(
        profile=profile,
        version=version,
        build_id=build_id,
        source_commit=source_commit,
        features=features,
        excluded_components=excluded_components,
        included_components=included_components,
        dependency_versions=dependency_versions,
    )
    if profile == "full":
        required_full_features = {"magic-segmentation"}
        if "area-models" not in excluded_components:
            required_full_features.add("area-inference")
        if not required_full_features.issubset(features):
            errors.append("full profile feature manifest is incomplete")
    if expected_profile is not None and profile != expected_profile:
        errors.append(f"profile mismatch: expected {expected_profile}, found {profile or '<empty>'}")
    if expected_version is not None and version != expected_version:
        errors.append(f"version mismatch: expected {expected_version}, found {version or '<empty>'}")
    if expected_commit is not None and source_commit != expected_commit:
        errors.append(f"stale dist: expected commit {expected_commit}, found {source_commit or '<empty>'}")
    if expected_excluded_components is not None:
        expected_exclusions = sorted(str(item).strip() for item in expected_excluded_components)
        if sorted(excluded_components) != expected_exclusions:
            errors.append(
                "excluded component mismatch: expected "
                f"{expected_exclusions}, found {sorted(excluded_components)}"
            )
    if expected_included_components is not None:
        expected_inclusions = sorted(str(item).strip() for item in expected_included_components)
        if sorted(included_components) != expected_inclusions:
            errors.append(
                "included component mismatch: expected "
                f"{expected_inclusions}, found {sorted(included_components)}"
            )
    if require_clean_source and bool(payload.get("source_dirty", True)):
        errors.append("release manifest records a dirty source tree")
    if require_clean_build and not bool(payload.get("clean_build", False)):
        errors.append("release manifest was produced by a non-clean build")

    marker_path = root / BUILD_ID_FILENAME
    try:
        marker = marker_path.read_text(encoding="utf-8").strip()
    except OSError as exc:
        errors.append(f"missing or unreadable {BUILD_ID_FILENAME}: {exc}")
    else:
        if not build_id or marker != build_id:
            errors.append("build-id mismatch between manifest and build-id.txt")

    entries = payload.get("files", [])
    if not isinstance(entries, list):
        errors.append("release manifest files must be a list")
        entries = []
    listed_paths: set[str] = set()
    for entry in entries:
        if not isinstance(entry, dict):
            errors.append("release manifest contains a non-object file entry")
            continue
        token = str(entry.get("path", "")).replace("\\", "/")
        if token in listed_paths:
            errors.append(f"duplicate manifest file entry: {token}")
            continue
        listed_paths.add(token)
        file_path = _safe_manifest_path(root, token)
        if file_path is None:
            errors.append(f"unsafe manifest file path: {token!r}")
            continue
        if not file_path.is_file():
            errors.append(f"missing packaged file: {token}")
            continue
        expected_size = entry.get("size")
        if not isinstance(expected_size, int) or file_path.stat().st_size != expected_size:
            errors.append(f"size mismatch: {token}")
            continue
        expected_hash = str(entry.get("sha256", ""))
        if len(expected_hash) != 64 or sha256_file(file_path) != expected_hash:
            errors.append(f"sha256 mismatch: {token}")
            continue
        result["checked_files"] = int(result["checked_files"]) + 1

    component_prefixes = {
        "area-models": "runtime/area-models/",
        "content-templates": "runtime/content-templates/",
    }
    overlap = sorted(set(excluded_components) & set(included_components))
    if overlap:
        errors.append("components cannot be both included and excluded: " + ", ".join(overlap))
    for component in excluded_components:
        prefix = component_prefixes.get(component)
        if prefix and any(path.startswith(prefix) for path in listed_paths):
            errors.append(f"excluded component is present in package inventory: {component}")
    for component in included_components:
        prefix = component_prefixes.get(component)
        if prefix and not any(path.startswith(prefix) for path in listed_paths):
            errors.append(f"included component is absent from package inventory: {component}")

    required_files = payload.get("required_runtime_files", [])
    if not isinstance(required_files, list):
        errors.append("required_runtime_files must be a list")
        required_files = []
    for required in required_files:
        token = str(required).replace("\\", "/")
        if token not in listed_paths:
            errors.append(f"required runtime asset is absent from manifest inventory: {token}")
        required_path = _safe_manifest_path(root, token)
        if required_path is None or not required_path.is_file():
            errors.append(f"missing required runtime asset: {token}")

    if reject_extra_files and root.is_dir():
        actual_paths = {
            path.relative_to(root).as_posix()
            for path in root.rglob("*")
            if path.is_file() and path.name != RELEASE_MANIFEST_FILENAME
        }
        extras = sorted(actual_paths - listed_paths)
        missing_inventory = sorted(listed_paths - actual_paths)
        if extras:
            errors.append("unmanifested packaged files: " + ", ".join(extras))
        if missing_inventory:
            errors.append("manifest inventory files missing from package: " + ", ".join(missing_inventory))

    result["ok"] = not errors
    return result


def run_release_self_check(app_root: str | Path | None = None) -> dict[str, Any]:
    root = Path(app_root or release_root())
    report = verify_release_manifest(root)
    functional_checks: dict[str, Any] = {}
    report["functional_checks"] = functional_checks
    if not report.get("ok"):
        return report

    errors = report.setdefault("errors", [])
    warnings = report.setdefault("warnings", [])
    for executable_name in ("FiberDiameterMeasurement.exe", "FiberAreaWorker.exe"):
        executable_path = root / executable_name
        valid, reason = _validate_pe_executable(executable_path)
        functional_checks[f"pe:{executable_name}"] = valid
        if not valid:
            errors.append(f"invalid Windows executable {executable_name}: {reason}")

    try:
        from fdm.models import Calibration

        probe = Calibration(
            mode="self_check",
            pixels_per_unit=5.0,
            unit="um",
            source_label="release-self-check",
        )
        measurement_ok = abs(probe.px_to_unit(25.0) - 5.0) <= 1e-12
    except Exception as exc:  # noqa: BLE001
        measurement_ok = False
        errors.append(f"core measurement self-check failed: {exc}")
    functional_checks["core_measurement"] = measurement_ok
    if not measurement_ok and not any("core measurement" in str(item) for item in errors):
        errors.append("core measurement self-check returned an unexpected value")

    try:
        from fdm.application_launch import SingleInstanceCoordinator
        from PySide6.QtNetwork import QLocalServer, QLocalSocket

        ipc_ok = all((SingleInstanceCoordinator, QLocalServer, QLocalSocket))
    except Exception as exc:  # noqa: BLE001
        ipc_ok = False
        errors.append(f"Qt local IPC self-check failed: {exc}")
    functional_checks["qt_local_ipc"] = ipc_ok
    if not ipc_ok and not any("Qt local IPC" in str(item) for item in errors):
        errors.append("Qt local IPC self-check is unavailable")

    features = set(str(item) for item in report.get("features", []))
    versions = report.get("dependency_versions", {})
    raster_probe: dict[str, Any] = {}
    if "image-export" in features:
        for distribution in ("Pillow", "tifffile"):
            version = str(versions.get(distribution, "")).strip()
            ok = bool(version and version != "not-installed")
            functional_checks[f"dependency:{distribution}"] = ok
            if not ok:
                errors.append(
                    "image-export dependency is unavailable: "
                    f"{distribution}"
                )
        try:
            raster_probe = _probe_pillow_raster_encoders()
        except Exception as exc:  # noqa: BLE001
            functional_checks["raster_encoders"] = False
            errors.append(f"Pillow raster encoder self-check failed: {exc}")
        else:
            functional_checks["raster_encoders"] = raster_probe
            if raster_probe.get("ok") is not True:
                errors.append("raster encoder production self-check returned a failure")
            format_results = raster_probe.get("formats", {})
            if not isinstance(format_results, dict):
                format_results = {}
            for format_name in ("png", "jpeg", "tiff", "bmp", "webp"):
                format_result = format_results.get(format_name, {})
                ok = (
                    isinstance(format_result, dict)
                    and format_result.get("ok") is True
                )
                functional_checks[f"raster_encoder:{format_name}"] = ok
                if not ok:
                    reason = (
                        str(format_result.get("message", "")).strip()
                        if isinstance(format_result, dict)
                        else ""
                    )
                    errors.append(
                        f"Pillow {format_name.upper()} encoder is unavailable"
                        + (f": {reason}" if reason else "")
                    )
            production_cases = raster_probe.get("production_cases", {})
            required_production_cases = (
                "gray16_png",
                "gray16_tiff_deflate",
                "gray16_tiff_lzw",
                "gray16_tiff_none",
                "gray32_float_tiff_deflate",
                "webp_lossy",
                "webp_lossless",
            )
            for case_name in required_production_cases:
                case_result = (
                    production_cases.get(case_name, {})
                    if isinstance(production_cases, dict)
                    else {}
                )
                ok = (
                    isinstance(case_result, dict)
                    and case_result.get("ok") is True
                )
                functional_checks[f"raster_production:{case_name}"] = ok
                if not ok:
                    reason = (
                        str(case_result.get("message", "")).strip()
                        if isinstance(case_result, dict)
                        else ""
                    )
                    errors.append(
                        f"raster production path self-check failed: {case_name}"
                        + (f": {reason}" if reason else "")
                    )
        try:
            tifffile_probe = _probe_tifffile_precision()
        except Exception as exc:  # noqa: BLE001
            functional_checks["tifffile_precision"] = False
            functional_checks["tifffile:uint16"] = False
            functional_checks["tifffile:float32"] = False
            errors.append(
                "tifffile 16-bit/float32 TIFF self-check failed: "
                f"{exc}"
            )
        else:
            functional_checks["tifffile_precision"] = tifffile_probe
            case_results = tifffile_probe.get("cases", {})
            for case_name in ("uint16", "float32"):
                case_result = (
                    case_results.get(case_name, {})
                    if isinstance(case_results, dict)
                    else {}
                )
                ok = (
                    isinstance(case_result, dict)
                    and case_result.get("ok") is True
                )
                functional_checks[f"tifffile:{case_name}"] = ok
                if not ok:
                    errors.append(
                        f"tifffile {case_name} TIFF round-trip is unavailable"
                    )

    if "image-processing" in features:
        try:
            processing_probe = _probe_image_processing_pipeline()
        except Exception as exc:  # noqa: BLE001
            functional_checks["image_processing_pipeline"] = False
            errors.append(f"image processing pipeline self-check failed: {exc}")
        else:
            processing_ok = processing_probe.get("ok") is True
            functional_checks["image_processing_pipeline"] = processing_probe
            if not processing_ok:
                errors.append("image processing pipeline self-check returned a failure")

    if {"image-analysis", "batch-processing"} & features:
        try:
            analysis_probe = _probe_analysis_pipeline()
        except Exception as exc:  # noqa: BLE001
            functional_checks["analysis_pipeline"] = False
            errors.append(f"analysis pipeline self-check failed: {exc}")
        else:
            analysis_ok = analysis_probe.get("ok") is True
            functional_checks["analysis_pipeline"] = analysis_probe
            if not analysis_ok:
                errors.append("analysis pipeline self-check returned a failure")

    if report.get("profile") == "full":
        for distribution in ("torch", "torchvision"):
            version = str(versions.get(distribution, "")).strip()
            ok = bool(version and version != "not-installed")
            functional_checks[f"dependency:{distribution}"] = ok
            if not ok:
                errors.append(
                    f"full profile dependency is unavailable: {distribution}"
                )

    execute_runtime_probe = bool(
        (sys.platform.startswith("win") and getattr(sys, "frozen", False))
        or os.environ.get("FDM_SELF_CHECK_EXECUTE", "").strip() == "1"
    )
    if "area-inference" in features and execute_runtime_probe and not errors:
        try:
            probe_result = _probe_area_worker(root)
        except Exception as exc:  # noqa: BLE001
            functional_checks["area_worker"] = False
            errors.append(f"area worker functional self-check failed: {exc}")
        else:
            functional_checks["area_worker"] = True
            functional_checks["area_worker_result"] = probe_result
    elif "area-inference" in features:
        warnings.append("Windows frozen area-worker execution probe skipped on this host")
        functional_checks["area_worker"] = "skipped_non_windows"
    elif "area-models" in report.get("excluded_components", []):
        functional_checks["area_worker"] = "skipped_area_models_excluded"

    report["ok"] = not errors
    return report


def _probe_pillow_raster_encoders() -> dict[str, Any]:
    """Exercise every format exposed by the raster export dialog.

    A PPM source keeps the probe independent from the PNG encoder under test.
    The production writer performs decode, encode, reopen/verify, fsync and
    atomic replacement, so this check also catches missing frozen plugins.
    """

    import numpy as np

    from fdm.services.raster_export import (
        RasterEncodingOptions,
        RasterExportFormat,
        RasterExportWriter,
    )
    from fdm.services.raster_io import (
        NativeTiffCompression,
        numpy_to_raster_plane,
        read_raster_file,
        write_native_raster_asset,
    )

    writer = RasterExportWriter()
    format_results: dict[str, dict[str, Any]] = {}
    production_cases: dict[str, dict[str, Any]] = {}
    backend_version = ""
    with tempfile.TemporaryDirectory(prefix="fdm-栅格编码-self-check-") as tmpdir:
        root = Path(tmpdir)
        source = root / "编码源.ppm"
        width, height = 4, 3
        pixels = bytes(
            channel
            for y in range(height)
            for x in range(width)
            for channel in (
                (x * 67 + y * 19) % 256,
                (x * 29 + y * 83) % 256,
                (x * 101 + y * 7) % 256,
            )
        )
        source.write_bytes(f"P6\n{width} {height}\n255\n".encode("ascii") + pixels)

        for export_format in RasterExportFormat:
            capability = writer.capability(export_format)
            if not backend_version:
                backend_version = capability.backend_version
            if not capability.available:
                format_results[export_format.value] = {
                    "ok": False,
                    "available": False,
                    "message": capability.reason,
                }
                continue
            target = root / f"编码结果{export_format.canonical_suffix}"
            result = writer.write_file(
                source,
                target,
                RasterEncodingOptions(format=export_format),
            )
            if result:
                format_results[export_format.value] = {
                    "ok": True,
                    "available": True,
                    "width": result.width,
                    "height": result.height,
                    "bytes": result.bytes_written,
                }
            else:
                failure = result.failure
                format_results[export_format.value] = {
                    "ok": False,
                    "available": True,
                    "code": failure.code.value if failure is not None else "",
                    "message": failure.message if failure is not None else "",
                    "detail": failure.detail if failure is not None else "",
                }

        gray16 = np.asarray(
            (
                (0, 1, 255, 256),
                (1024, 32768, 65534, 65535),
                (7, 4095, 50000, 42),
            ),
            dtype=np.uint16,
        )
        gray32 = np.asarray(
            (
                (-12.5, -0.0, 0.0, 0.125),
                (1.0 / 3.0, 1.5, 65535.25, 1.0e8),
                (-1.0e-6, 42.75, -32768.5, 9.0),
            ),
            dtype=np.float32,
        )
        native_cases = (
            (
                "gray16_png",
                numpy_to_raster_plane(gray16),
                root / "16位灰度生产路径.png",
                {},
            ),
            (
                "gray16_tiff_deflate",
                numpy_to_raster_plane(gray16),
                root / "16位灰度-Deflate.tif",
                {"tiff_compression": NativeTiffCompression.DEFLATE},
            ),
            (
                "gray16_tiff_lzw",
                numpy_to_raster_plane(gray16),
                root / "16位灰度-LZW.tif",
                {"tiff_compression": NativeTiffCompression.LZW},
            ),
            (
                "gray16_tiff_none",
                numpy_to_raster_plane(gray16),
                root / "16位灰度-无压缩.tif",
                {"tiff_compression": NativeTiffCompression.NONE},
            ),
            (
                "gray32_float_tiff_deflate",
                numpy_to_raster_plane(gray32),
                root / "32位浮点-Deflate.tif",
                {"tiff_compression": NativeTiffCompression.DEFLATE},
            ),
        )
        for case_name, plane, target, options in native_cases:
            write_result = write_native_raster_asset(plane, target, **options)
            read_result = read_raster_file(target) if write_result else None
            restored = None if read_result is None else read_result.plane
            ok = bool(
                write_result
                and read_result
                and restored is not None
                and restored.pixel_type is plane.pixel_type
                and restored.width == plane.width
                and restored.height == plane.height
                and restored.sha256() == plane.sha256()
            )
            production_cases[case_name] = {
                "ok": ok,
                "pixel_type": plane.pixel_type.value,
                "width": plane.width,
                "height": plane.height,
                "bytes": int(write_result.bytes_written),
                "sha256": plane.sha256(),
                "message": (
                    ""
                    if ok
                    else (
                        write_result.failure.message
                        if write_result.failure is not None
                        else (
                            read_result.failure.message
                            if read_result is not None
                            and read_result.failure is not None
                            else "像素类型、尺寸或 SHA256 不一致"
                        )
                    )
                ),
            }

        rgb_array = np.frombuffer(pixels, dtype=np.uint8).reshape(
            height,
            width,
            3,
        )
        rgb_plane = numpy_to_raster_plane(rgb_array)
        for case_name, lossless, quality in (
            ("webp_lossy", False, 73),
            ("webp_lossless", True, 100),
        ):
            target = root / f"{case_name}-中文.webp"
            result = writer.write_file(
                source,
                target,
                RasterEncodingOptions(
                    format=RasterExportFormat.WEBP,
                    quality=quality,
                    webp_lossless=lossless,
                ),
            )
            restored_result = read_raster_file(target) if result else None
            restored = (
                None
                if restored_result is None
                else restored_result.plane
            )
            exact_ok = (
                restored is not None
                and restored.sha256() == rgb_plane.sha256()
            )
            ok = bool(
                result
                and restored_result
                and restored is not None
                and restored.width == width
                and restored.height == height
                and (not lossless or exact_ok)
            )
            production_cases[case_name] = {
                "ok": ok,
                "lossless": lossless,
                "quality": quality,
                "width": int(result.width),
                "height": int(result.height),
                "bytes": int(result.bytes_written),
                "exact_sha256": bool(exact_ok),
                "message": (
                    ""
                    if ok
                    else (
                        result.failure.message
                        if result.failure is not None
                        else "WebP 解码尺寸或无损像素 SHA256 不一致"
                    )
                ),
            }

    return {
        "ok": all(
            result.get("ok") is True
            for result in format_results.values()
        )
        and len(format_results) == len(RasterExportFormat)
        and all(
            result.get("ok") is True
            for result in production_cases.values()
        ),
        "backend": "Pillow",
        "backend_version": backend_version,
        "formats": format_results,
        "production_cases": production_cases,
    }


def _probe_image_processing_pipeline() -> dict[str, Any]:
    """Exercise OpenCV processing and authoritative native-asset round trips."""

    import numpy as np

    from fdm.services.image_processing import (
        ImageOperation,
        ImageOperationRequest,
        execute_image_operation,
    )
    from fdm.services.raster_io import (
        numpy_to_raster_plane,
        read_raster_file,
        write_native_raster_asset,
    )

    sources = {
        "gray8": np.arange(48, dtype=np.uint8).reshape(6, 8),
        "gray16": (
            np.arange(48, dtype=np.uint16).reshape(6, 8) * 1201
        ),
        "gray32_float": (
            np.arange(48, dtype=np.float32).reshape(6, 8) / 7.0 - 2.0
        ),
    }
    cases: dict[str, dict[str, Any]] = {}
    with tempfile.TemporaryDirectory(
        prefix="fdm-图像处理管线-self-check-"
    ) as tmpdir:
        root = Path(tmpdir)
        for index, (case_name, source) in enumerate(sources.items(), start=1):
            request_id = f"self-check-processing-{case_name}"
            generation = 40 + index
            result = execute_image_operation(
                ImageOperationRequest.create(
                    ImageOperation.GAUSSIAN_BLUR,
                    source,
                    request_id=request_id,
                    generation=generation,
                    kernel_size=3,
                    sigma=0.8,
                )
            )
            if (
                result.request_id != request_id
                or result.generation != generation
            ):
                raise RuntimeError(
                    f"{case_name} processing request identity changed"
                )
            if result.image.shape != source.shape or result.image.dtype != source.dtype:
                raise RuntimeError(
                    f"{case_name} processing changed dtype or dimensions"
                )
            plane = numpy_to_raster_plane(result.image)
            suffix = ".tif" if case_name == "gray32_float" else ".png"
            target = root / f"{case_name}-处理结果-中文{suffix}"
            written = write_native_raster_asset(plane, target)
            if not written:
                failure = written.failure
                raise RuntimeError(
                    failure.message if failure is not None else "native write failed"
                )
            restored_result = read_raster_file(target)
            restored = restored_result.plane
            if (
                not restored_result
                or restored is None
                or restored.pixel_type is not plane.pixel_type
                or restored.width != plane.width
                or restored.height != plane.height
                or restored.sha256() != plane.sha256()
            ):
                raise RuntimeError(
                    f"{case_name} native asset round-trip changed pixels"
                )
            cases[case_name] = {
                "ok": True,
                "request_id": result.request_id,
                "generation": result.generation,
                "pixel_type": plane.pixel_type.value,
                "width": plane.width,
                "height": plane.height,
                "sha256": plane.sha256(),
                "bytes": written.bytes_written,
            }
    return {
        "ok": len(cases) == len(sources)
        and all(item["ok"] is True for item in cases.values()),
        "operation": ImageOperation.GAUSSIAN_BLUR.value,
        "cases": cases,
    }


def _probe_analysis_pipeline() -> dict[str, Any]:
    """Exercise safe assets, advanced registry, workbook and batch contracts."""

    import numpy as np
    from openpyxl import load_workbook

    from fdm.analysis_artifacts import AnalysisArtifact, AnalysisTable
    from fdm.image_processing_models import (
        ImageOperationSpec,
        ImageProcessingRecipe,
    )
    from fdm.services.advanced_analysis_registry import (
        AdvancedAnalysisInvocation,
        AdvancedAnalysisRegistry,
    )
    from fdm.services.advanced_image_analysis import AdvancedAnalysisKind
    from fdm.services.analysis_asset_io import (
        inspect_safe_analysis_npz,
        write_safe_analysis_npz,
    )
    from fdm.services.analysis_export import export_analysis_workbook
    from fdm.services.image_batch import (
        BatchRasterInput,
        BatchRecipeRequest,
        execute_batch_recipe,
    )
    from fdm.services.raster_io import numpy_to_raster_plane

    with tempfile.TemporaryDirectory(
        prefix="fdm-分析管线-self-check-"
    ) as tmpdir:
        root = Path(tmpdir)
        asset_path = root / "分析资产" / "安全数组.npz"
        asset_info = write_safe_analysis_npz(
            asset_path,
            schema="fdm.self-check.analysis.v1",
            arrays={
                "values": np.asarray([[1, 2], [3, 4]], dtype=np.uint16),
                "mask": np.asarray([[True, False], [False, True]], dtype=bool),
            },
            metadata={"来源": "发布自检", "request_id": "analysis-self-check"},
        )
        inspected = inspect_safe_analysis_npz(asset_path)
        if (
            inspected.sha256 != asset_info.sha256
            or inspected.schema != "fdm.self-check.analysis.v1"
            or {item[0] for item in inspected.members} != {"mask", "values"}
        ):
            raise RuntimeError("safe analysis NPZ verification failed")

        artifact = AnalysisArtifact(
            id="analysis-self-check",
            source_document_id="中文来源图片",
            source_pixel_revision=1,
            tool_id="fdm.histogram",
            tool_version="1",
            parameters={"bins": 4, "channel": "亮度"},
            scalars={"有效 N": 4, "均值": 2.5},
            tables=(
                AnalysisTable(
                    name="自检明细",
                    columns=("序号", "值"),
                    rows=((1, 1.0), (2, 4.0)),
                ),
            ),
            created_at="2026-07-27T00:00:00+00:00",
        )
        workbook_path = root / "中文目录" / "分析结果自检.xlsx"
        workbook_result = export_analysis_workbook(
            (artifact,),
            workbook_path,
            document_names={"中文来源图片": "中文来源图片"},
        )
        if not workbook_result.success or workbook_result.path is None:
            raise RuntimeError(workbook_result.message)
        workbook = load_workbook(workbook_result.path, read_only=True, data_only=False)
        try:
            required_sheets = {"分析摘要", "参数与来源"}
            if not required_sheets.issubset(workbook.sheetnames):
                raise RuntimeError("analysis workbook is missing required sheets")
            if workbook["分析摘要"]["A1"].value != "分析结果ID":
                raise RuntimeError("analysis workbook summary header is invalid")
            workbook_sheets = list(workbook.sheetnames)
        finally:
            workbook.close()

        direction_image = np.zeros((32, 40), dtype=np.uint8)
        direction_image[14:18, 3:37] = 255
        direction_plane = numpy_to_raster_plane(direction_image)
        registry = AdvancedAnalysisRegistry()
        advanced = registry.execute(
            AdvancedAnalysisInvocation(
                AdvancedAnalysisKind.DIRECTIONALITY,
                request_id="analysis-advanced-self-check",
                generation=7,
                plane=direction_plane,
                parameters={"bins": 18, "channel": "luminance"},
            )
        )
        if (
            advanced.request_id != "analysis-advanced-self-check"
            or advanced.generation != 7
            or float(advanced.result.total_weight) <= 0.0
        ):
            raise RuntimeError("advanced analysis registry contract failed")

        recipe = ImageProcessingRecipe.from_operations(
            (
                ImageOperationSpec("add", {"value": 2.0}),
                ImageOperationSpec("invert", {}),
            )
        )
        batch_request = BatchRecipeRequest(
            request_id="analysis-batch-self-check",
            generation=11,
            recipe=recipe,
            inputs=(
                BatchRasterInput(
                    document_id="batch-one",
                    display_name="批处理一",
                    raster=numpy_to_raster_plane(
                        np.arange(16, dtype=np.uint8).reshape(4, 4)
                    ),
                    source_pixel_revision=1,
                ),
                BatchRasterInput(
                    document_id="batch-two",
                    display_name="批处理二",
                    raster=numpy_to_raster_plane(
                        np.arange(16, dtype=np.uint16).reshape(4, 4)
                    ),
                    source_pixel_revision=2,
                ),
            ),
            available_disk_bytes=10 << 30,
        )
        batch_result = execute_batch_recipe(batch_request)
        if (
            batch_result.request_id != batch_request.request_id
            or batch_result.generation != batch_request.generation
            or batch_result.success_count != 2
            or len(batch_result.commit_candidates) != 2
            or "成功 2 张" not in batch_result.summary_text
        ):
            raise RuntimeError("batch request/generation/summary contract failed")
        for candidate in batch_result.commit_candidates:
            if (
                candidate.raster.sha256()
                != candidate.derivation.result_sha256
            ):
                raise RuntimeError("batch result digest contract failed")

        return {
            "ok": True,
            "safe_npz": {
                "ok": True,
                "schema": inspected.schema,
                "members": [item[0] for item in inspected.members],
                "sha256": inspected.sha256,
            },
            "workbook": {
                "ok": True,
                "sheets": workbook_sheets,
                "bytes": workbook_result.path.stat().st_size,
                "unicode_path": True,
            },
            "advanced_analysis": {
                "ok": True,
                "kind": advanced.kind.value,
                "request_id": advanced.request_id,
                "generation": advanced.generation,
                "registered_tools": len(registry.registrations()),
            },
            "batch": {
                "ok": True,
                "request_id": batch_result.request_id,
                "generation": batch_result.generation,
                "item_count": len(batch_result.items),
                "success_count": batch_result.success_count,
                "summary": batch_result.summary_text,
                "result_sha256": [
                    candidate.raster.sha256()
                    for candidate in batch_result.commit_candidates
                ],
            },
        }


def _probe_tifffile_precision() -> dict[str, Any]:
    """Verify that the frozen runtime preserves 16-bit and float32 samples."""

    import numpy as np
    import tifffile

    cases = {
        "uint16": np.asarray(
            [
                [0, 1, 255, 256],
                [1024, 32768, 65534, 65535],
                [7, 4095, 50000, 42],
            ],
            dtype=np.uint16,
        ),
        "float32": np.asarray(
            [
                [-12.5, -0.0, 0.0, 0.125],
                [1.0 / 3.0, 1.5, 65535.25, 1.0e8],
                [-1.0e-6, 42.75, -32768.5, 9.0],
            ],
            dtype=np.float32,
        ),
    }
    results: dict[str, dict[str, Any]] = {}
    with tempfile.TemporaryDirectory(prefix="fdm-TIFF位深-self-check-") as tmpdir:
        root = Path(tmpdir)
        for case_name, expected in cases.items():
            path = root / f"{case_name}-中文路径.tif"
            tifffile.imwrite(
                path,
                expected,
                photometric="minisblack",
                metadata=None,
                compression=None,
            )
            actual = np.asarray(tifffile.imread(path))
            dtype_ok = actual.dtype == expected.dtype
            shape_ok = actual.shape == expected.shape
            samples_ok = (
                dtype_ok
                and shape_ok
                and actual.tobytes(order="C") == expected.tobytes(order="C")
            )
            result = {
                "ok": bool(dtype_ok and shape_ok and samples_ok),
                "dtype": str(actual.dtype),
                "shape": [int(value) for value in actual.shape],
                "bytes": int(path.stat().st_size),
            }
            results[case_name] = result
            if not result["ok"]:
                raise RuntimeError(
                    f"{case_name} round-trip changed dtype, shape or sample bits"
                )
    return {
        "ok": all(result["ok"] for result in results.values()),
        "backend": "tifffile",
        "backend_version": str(getattr(tifffile, "__version__", "")),
        "cases": results,
    }


def _validate_pe_executable(path: Path) -> tuple[bool, str]:
    try:
        with path.open("rb") as stream:
            dos_header = stream.read(64)
            if len(dos_header) < 64 or dos_header[:2] != b"MZ":
                return False, "missing DOS MZ header"
            pe_offset = int.from_bytes(dos_header[60:64], "little")
            if pe_offset < 64 or pe_offset > 16 * 1024 * 1024:
                return False, "invalid PE header offset"
            stream.seek(pe_offset)
            if stream.read(4) != b"PE\x00\x00":
                return False, "missing PE signature"
    except OSError as exc:
        return False, str(exc)
    return True, ""


def _probe_area_worker(root: Path) -> dict[str, Any]:
    worker = root / "FiberAreaWorker.exe"
    weights_dir = root / "runtime" / "area-models"
    vendor_root = root / "runtime" / "area-infer" / "vendor" / "yolact"
    model_files = sorted(weights_dir.glob("*.pth"))
    if not model_files:
        raise RuntimeError("no packaged area model")
    preferred_model = weights_dir / "b_v1_1.3.pth"
    model_file = preferred_model if preferred_model.is_file() else model_files[0]
    request_id = "self-check-infer"
    with tempfile.TemporaryDirectory(prefix="fdm-面积识别-self-check-") as tmpdir:
        from PIL import Image

        image_path = Path(tmpdir) / "面积识别自检.png"
        Image.new("RGB", (1280, 960), color=(255, 255, 255)).save(image_path, format="PNG")
        request = {
            "protocol": AREA_WORKER_PROTOCOL,
            "version": AREA_WORKER_PROTOCOL_VERSION,
            "request_id": request_id,
            "op": "infer",
            "image": {"path": str(image_path)},
            "model": {"name": model_file.stem, "file": model_file.name},
            "runtime": {
                "weights_dir": str(weights_dir),
                "vendor_root": str(vendor_root),
                "device": "cpu",
                "require_trusted_weights": True,
                "verify_trusted_weights": True,
            },
            "options": {
                "include_overlay": False,
                "inference": {"top_k": 5, "nms_top_k": 20},
            },
        }
        payload = json.dumps(request, ensure_ascii=True, allow_nan=False)
        completed = subprocess.run(
            [str(worker)],
            input=payload,
            text=True,
            encoding="utf-8",
            errors="replace",
            capture_output=True,
            timeout=300,
            check=False,
        )
    lines = [line for line in completed.stdout.splitlines() if line.strip()]
    if completed.returncode != 0 or len(lines) != 1:
        raise RuntimeError(
            f"worker protocol failed rc={completed.returncode}, lines={len(lines)}, "
            f"stderr={completed.stderr[-1000:]}"
        )
    response = json.loads(lines[0])
    if response.get("protocol") != AREA_WORKER_PROTOCOL:
        raise RuntimeError("worker returned an invalid protocol envelope")
    if response.get("request_id") != request_id or response.get("ok") is not True:
        raise RuntimeError(f"worker CPU inference failed: {response.get('error')}")
    result = response.get("result")
    if not isinstance(result, dict) or not isinstance(result.get("instances"), list):
        raise RuntimeError("worker inference result schema is invalid")
    return {
        "mode": "one_shot",
        "device": "cpu",
        "model_file": model_file.name,
        "image_size": [1280, 960],
        "unicode_path": True,
        "instance_count": len(result["instances"]),
    }


def packaged_runtime_features(app_root: str | Path | None = None) -> frozenset[str]:
    """Return the explicit packaged feature manifest; source runs expose all features."""

    root = Path(app_root) if app_root is not None else release_root()
    manifest_path = root / RELEASE_MANIFEST_FILENAME
    if not manifest_path.is_file():
        features = {"measurement", "capture", "digital-slide"}
        runtime_root = root / "runtime"
        if (runtime_root / "area-infer").is_dir() and (runtime_root / "area-models").is_dir():
            features.add("area-inference")
        if (runtime_root / "segment-anything").is_dir():
            features.add("magic-segmentation")
        return frozenset(features)
    try:
        payload = json.loads(manifest_path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return frozenset()
    features = payload.get("features", []) if isinstance(payload, dict) else []
    if not isinstance(features, list):
        return frozenset()
    return frozenset(str(item).strip() for item in features if str(item).strip())


def runtime_capability_hint(app_root: str | Path | None = None) -> str:
    root = Path(app_root) if app_root is not None else release_root()
    manifest_path = root / RELEASE_MANIFEST_FILENAME
    features = packaged_runtime_features(root)
    missing = []
    if "area-inference" not in features:
        missing.append("面积推理")
    if "magic-segmentation" not in features:
        missing.append("智能分割")
    if not missing:
        return ""
    if manifest_path.is_file():
        return f"当前发布 profile 未包含：{'、'.join(missing)}。相关入口已隐藏。"
    return (
        f"当前开发 wheel 未包含完整运行时资产：{'、'.join(missing)}。"
        "正式终端用户交付请使用 Windows full 安装包。"
    )


def format_self_check_report(report: dict[str, Any]) -> str:
    lines = [
        "Release self-check: " + ("PASS" if report.get("ok") else "FAIL"),
        f"Root: {report.get('root', '')}",
        f"Version: {report.get('version', '')}",
        f"Profile: {report.get('profile', '')}",
        f"Build ID: {report.get('build_id', '')}",
        f"Checked files: {report.get('checked_files', 0)}",
    ]
    for error in report.get("errors", []):
        lines.append(f"ERROR: {error}")
    for warning in report.get("warnings", []):
        lines.append(f"WARNING: {warning}")
    return "\n".join(lines)
