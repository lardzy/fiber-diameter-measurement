from __future__ import annotations

import csv
import hashlib
from io import StringIO
import json
from pathlib import Path
import zipfile

import numpy as np
from openpyxl import load_workbook

from fdm.analysis_artifacts import (
    AnalysisArtifact,
    AnalysisArtifactStatus,
    AnalysisAssetKind,
    AnalysisAssetReference,
    AnalysisCurve,
    AnalysisObjectKind,
    AnalysisObjectReference,
    AnalysisTable,
)
from fdm.services.analysis_asset_io import write_safe_analysis_npz
from fdm.services.analysis_export import (
    AnalysisExportFailureCode,
    AnalysisExportService,
)
from fdm.services.analysis_profiles import ANALYSIS_OUTPUT_FIELDS_PARAMETER


def _artifact(
    *,
    artifact_id: str = "analysis_1",
    tool_id: str = "fdm.histogram",
    status: AnalysisArtifactStatus = AnalysisArtifactStatus.CURRENT,
    assets: tuple[AnalysisAssetReference, ...] = (),
) -> AnalysisArtifact:
    return AnalysisArtifact(
        id=artifact_id,
        source_document_id="doc_1",
        source_pixel_revision=2,
        source_reference=AnalysisObjectReference(
            AnalysisObjectKind.ROI,
            "roi_1",
            3,
        ),
        tool_id=tool_id,
        tool_version="1.2",
        parameters={"bins": 64, "channel": "亮度", "category_label": "玻璃纤维"},
        calibration_signature="sha256:" + ("a" * 64),
        scalars={"有效 N": 4, "均值": 2.5},
        tables=(
            AnalysisTable(
                name="直方图明细",
                columns=("灰度", "频数", "说明"),
                rows=((0.0, 1, "低值"), (1.0, 3, "=不作为公式")),
            ),
        ),
        curves=(
            AnalysisCurve(
                name="直方图",
                x=(0.0, 1.0),
                y=(1.0, 3.0),
                x_unit="灰度",
                y_unit="频数",
            ),
        ),
        assets=assets,
        status=status,
        stale_reason="来源 ROI 已变化" if status is AnalysisArtifactStatus.STALE else None,
        created_at="2026-07-27T08:00:00+00:00",
    )


def test_analysis_workbook_uses_independent_chinese_pages(tmp_path: Path) -> None:
    service = AnalysisExportService()
    current = _artifact()
    stale = _artifact(
        artifact_id="analysis_2",
        tool_id="fdm.particle_analysis",
        status=AnalysisArtifactStatus.STALE,
    )

    result = service.export_workbook(
        (current, stale),
        tmp_path / "公司分析结果.anything",
        document_names={"doc_1": "25A 显微图片"},
        roi_names={"roi_1": "孔洞区域"},
    )

    assert result.success
    assert result.failure_code is None
    assert result.path == tmp_path / "公司分析结果.xlsx"
    assert result.artifact_count == 2
    assert result.sheet_count == 4
    workbook = load_workbook(result.path, data_only=False)
    try:
        assert workbook.sheetnames[:2] == ["分析摘要", "参数与来源"]
        assert workbook["分析摘要"]["A1"].value == "分析结果ID"
        assert workbook["分析摘要"]["B1"].value == "状态"
        assert workbook["分析摘要"]["D2"].value == "25A 显微图片"
        assert "ROI：孔洞区域" in workbook["分析摘要"]["F2"].value
        assert workbook["分析摘要"]["B3"].value == "已失效"
        assert workbook["分析摘要"]["C3"].value == "来源 ROI 已变化"
        detail_names = workbook.sheetnames[2:]
        assert any("直方图" in name for name in detail_names)
        assert any("粒子分析" in name for name in detail_names)
        histogram_sheet = workbook[next(name for name in detail_names if "直方图" in name)]
        values = [
            cell.value
            for row in histogram_sheet.iter_rows()
            for cell in row
        ]
        assert "直方图明细" in values
        assert "'=不作为公式" in values
    finally:
        workbook.close()


def test_workbook_preserves_output_selection_and_analysis_warnings(
    tmp_path: Path,
) -> None:
    artifact = AnalysisArtifact(
        id="selected-intensity",
        source_document_id="doc_1",
        source_pixel_revision=2,
        tool_id="fdm.intensity",
        tool_version="2",
        parameters={
            "channel": "luminance",
            ANALYSIS_OUTPUT_FIELDS_PARAMETER: [
                "central_tendency",
                "percentiles",
            ],
        },
        scalars={"valid_pixel_count": 4, "mean": 2.5},
        warnings=("阈值范围内没有像素。",),
    )

    result = AnalysisExportService().export_workbook(
        (artifact,),
        tmp_path / "字段选择.xlsx",
    )

    assert result.success
    workbook = load_workbook(result.path, data_only=False)
    try:
        rows = tuple(workbook["参数与来源"].iter_rows(values_only=True))
        assert any(
            row[2:] == (
                "参数",
                "输出字段选择",
                "均值、中位数与众数、分位数",
            )
            for row in rows
        )
        assert any(
            row[2:] == ("提示", "分析提示", "阈值范围内没有像素。")
            for row in rows
        )
    finally:
        workbook.close()


def test_analysis_table_csv_is_utf8_bom_chinese_and_not_a_formula(tmp_path: Path) -> None:
    artifact = _artifact()

    result = AnalysisExportService().export_table_csv(
        artifact,
        "直方图明细",
        tmp_path / "直方图.txt",
    )

    assert result.success
    assert result.path == tmp_path / "直方图.csv"
    payload = result.path.read_bytes()
    assert payload.startswith(b"\xef\xbb\xbf")
    rows = list(csv.reader(StringIO(payload.decode("utf-8-sig"))))
    assert rows[0] == ["灰度", "频数", "说明"]
    assert rows[1] == ["0.0", "1", "低值"]
    assert rows[2][2] == "'=不作为公式"


def test_analysis_curve_csv_preserves_units_and_missing_values(tmp_path: Path) -> None:
    artifact = _artifact()

    result = AnalysisExportService().export_curve_csv(
        artifact,
        "直方图",
        tmp_path / "curve.any",
    )

    assert result.success
    rows = list(
        csv.reader(
            StringIO(result.path.read_bytes().decode("utf-8-sig"))
        )
    )
    assert rows[0] == ["X (灰度)", "Y (频数)"]
    assert rows[1] == ["0.0", "1.0"]


def test_missing_artifacts_and_table_return_structured_failures(tmp_path: Path) -> None:
    service = AnalysisExportService()

    empty = service.export_workbook((), tmp_path / "empty.xlsx")
    invalid = service.export_workbook((_artifact(),), "")
    missing = service.export_table_csv(
        _artifact(),
        "不存在",
        tmp_path / "missing.csv",
    )

    assert not empty.success
    assert empty.failure_code is AnalysisExportFailureCode.EMPTY_SELECTION
    assert not invalid.success
    assert invalid.failure_code is AnalysisExportFailureCode.INVALID_TARGET
    assert not missing.success
    assert missing.failure_code is AnalysisExportFailureCode.TABLE_NOT_FOUND
    assert not (tmp_path / "missing.csv").exists()


def test_atomic_write_failure_preserves_existing_workbook(
    tmp_path: Path,
    monkeypatch,
) -> None:
    target = tmp_path / "analysis.xlsx"
    target.write_bytes(b"old-project-byte-content")

    def fail_write(path, payload) -> None:
        assert Path(path) == target
        assert payload
        raise OSError("injected write failure")

    monkeypatch.setattr(
        "fdm.services.analysis_export.atomic_write_bytes",
        fail_write,
    )

    result = AnalysisExportService().export_workbook((_artifact(),), target)

    assert not result.success
    assert result.failure_code is AnalysisExportFailureCode.ATOMIC_COMMIT_FAILED
    assert target.read_bytes() == b"old-project-byte-content"


def test_portable_package_contains_manifest_workbook_csv_and_verified_assets(
    tmp_path: Path,
) -> None:
    relative = "analysis/source/labels.npz"
    source = tmp_path / relative
    info = write_safe_analysis_npz(
        source,
        schema="fdm.test-labels.v1",
        arrays={"labels": np.asarray([[0, 1], [1, 0]], dtype=np.int32)},
    )
    reference = AnalysisAssetReference(
        kind=AnalysisAssetKind.LABEL_IMAGE,
        path=relative,
        sha256=info.sha256,
        media_type="application/x-npz",
        metadata={
            "schema": info.schema,
            "allow_pickle": False,
            "members": {
                name: {"dtype": dtype, "shape": list(shape)}
                for name, dtype, shape in info.members
            },
        },
    )
    artifact = _artifact(assets=(reference,))

    result = AnalysisExportService().export_portable_package(
        (artifact,),
        tmp_path / "portable.any",
        asset_root=tmp_path,
    )

    assert result.success
    assert result.path == tmp_path / "portable.zip"
    with zipfile.ZipFile(result.path) as archive:
        names = archive.namelist()
        assert "manifest.json" in names
        assert "analysis-results.xlsx" in names
        assert any(name.startswith("csv/") and "table-" in name for name in names)
        assert any(name.startswith("csv/") and "curve-" in name for name in names)
        manifest = json.loads(archive.read("manifest.json"))
        assert manifest["schema_version"] == 1
        assert manifest["artifacts"][0]["id"] == artifact.id
        asset_record = manifest["assets"][0]
        asset_payload = archive.read(asset_record["package_path"])
        assert hashlib.sha256(asset_payload).hexdigest() == reference.sha256
        for record in manifest["files"]:
            payload = archive.read(record["path"])
            assert len(payload) == record["size"]
            assert hashlib.sha256(payload).hexdigest() == record["sha256"]


def test_portable_package_rejects_missing_or_hash_mismatched_asset(
    tmp_path: Path,
) -> None:
    reference = AnalysisAssetReference(
        kind=AnalysisAssetKind.OTHER,
        path="analysis/missing/data.bin",
        sha256="a" * 64,
        media_type="application/octet-stream",
        metadata={"schema": "fdm.test-data.v1"},
    )
    missing = AnalysisExportService().export_portable_package(
        (_artifact(assets=(reference,)),),
        tmp_path / "missing.zip",
        asset_root=tmp_path,
    )
    source = tmp_path / reference.path
    source.parent.mkdir(parents=True, exist_ok=True)
    source.write_bytes(b"wrong")
    mismatch = AnalysisExportService().export_portable_package(
        (_artifact(assets=(reference,)),),
        tmp_path / "mismatch.zip",
        asset_root=tmp_path,
    )

    assert missing.failure_code is AnalysisExportFailureCode.ASSET_NOT_FOUND
    assert mismatch.failure_code is AnalysisExportFailureCode.ASSET_HASH_MISMATCH
