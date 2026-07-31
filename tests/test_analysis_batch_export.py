from __future__ import annotations

from pathlib import Path

import numpy as np
from openpyxl import load_workbook

from fdm.services.advanced_analysis_registry import AdvancedAnalysisInvocation
from fdm.services.analysis_batch import (
    AnalysisBatchItemResult,
    AnalysisBatchRequest,
    AnalysisBatchResult,
    AnalysisInvocation,
    builtin_plane_analysis_recipes,
    execute_analysis_batch,
)
from fdm.services.analysis_export import (
    AnalysisExportFailureCode,
    AnalysisExportService,
    build_analysis_batch_export_rows,
    export_analysis_batch_workbook,
)
from fdm.services.raster_io import numpy_to_raster_plane


def _batch_result_with_failure() -> AnalysisBatchResult:
    recipe = builtin_plane_analysis_recipes()[0]
    rows, columns = np.indices((32, 32))
    image = ((rows * 13 + columns * 7) % 256).astype(np.uint8)
    plane = numpy_to_raster_plane(image)
    roi_mask = columns < 16
    invocations = (
        AnalysisInvocation(
            item_id="doc_1",
            display_name="图像一",
            analysis=AdvancedAnalysisInvocation(
                recipe.kind,
                request_id="batch-export:doc_1",
                generation=2,
                plane=plane,
            ),
        ),
        AnalysisInvocation(
            item_id="doc_1::roi::roi_1",
            display_name="图像一 · ROI：边缘区域",
            analysis=AdvancedAnalysisInvocation(
                recipe.kind,
                request_id="batch-export:doc_1:roi_1",
                generation=2,
                plane=plane,
                roi_mask=roi_mask,
            ),
        ),
    )
    completed = execute_analysis_batch(
        AnalysisBatchRequest(
            request_id="batch-export",
            generation=2,
            recipe=recipe,
            invocations=invocations,
        )
    )
    return AnalysisBatchResult(
        request_id=completed.request_id,
        generation=completed.generation,
        recipe_id=completed.recipe_id,
        item_results=completed.item_results
        + (
            AnalysisBatchItemResult(
                item_id="doc_2::roi::roi_2",
                display_name="图像二 · ROI：失败区域",
                success=False,
                error_type="ValueError",
                error_message="=危险公式",
            ),
        ),
    )


def _multi_tool_batch_result() -> AnalysisBatchResult:
    recipe = builtin_plane_analysis_recipes()[-1]
    rows, columns = np.indices((32, 32))
    plane = numpy_to_raster_plane(
        ((rows * 5 + columns * 11) % 256).astype(np.uint8)
    )
    return execute_analysis_batch(
        AnalysisBatchRequest(
            request_id="batch-multi-export",
            generation=3,
            recipe=recipe,
            invocations=(
                AnalysisInvocation(
                    item_id="doc_multi",
                    display_name="多工具图像",
                    analysis=AdvancedAnalysisInvocation(
                        recipe.kind,
                        request_id="batch-multi-export:doc_multi",
                        generation=3,
                        plane=plane,
                    ),
                ),
            ),
        )
    )


def _header_indexes(sheet) -> dict[str, int]:
    return {
        str(cell.value): index
        for index, cell in enumerate(sheet[1], start=1)
    }


def test_batch_workbook_has_four_auditable_typed_sheets(
    tmp_path: Path,
) -> None:
    result = export_analysis_batch_workbook(
        _batch_result_with_failure(),
        tmp_path / "批量结果.anything",
        document_names={"doc_1": "样品 A", "doc_2": "=样品 B"},
        roi_names={"roi_1": "边缘 ROI", "roi_2": "失败 ROI"},
    )

    assert result.success
    assert result.failure_code is None
    assert result.path == tmp_path / "批量结果.xlsx"
    assert (result.item_count, result.success_count, result.failure_count) == (
        3,
        2,
        1,
    )
    assert result.sheet_count == 4

    workbook = load_workbook(result.path, data_only=False)
    try:
        assert workbook.sheetnames == ["总览", "逐图片", "逐 ROI", "失败明细"]
        overview = workbook["总览"]
        assert overview["A2"].value == "batch-export"
        assert overview["D2"].value == "完成"
        assert tuple(overview.cell(2, column).value for column in range(5, 10)) == (
            3,
            2,
            1,
            1,
            2,
        )

        by_image = workbook["逐图片"]
        image_headers = _header_indexes(by_image)
        assert by_image.max_row == 2
        assert by_image.cell(2, image_headers["来源图片"]).value == "样品 A"
        assert by_image.cell(2, image_headers["状态"]).value == "成功"
        assert isinstance(
            by_image.cell(2, image_headers["指标·valid_gradient_pixels"]).value,
            int,
        )
        assert isinstance(
            by_image.cell(2, image_headers["指标·concentration"]).value,
            float,
        )
        assert by_image.cell(2, image_headers["数组数量"]).value == 3

        by_roi = workbook["逐 ROI"]
        roi_headers = _header_indexes(by_roi)
        assert by_roi.max_row == 3
        assert by_roi.cell(2, roi_headers["ROI"]).value == "边缘 ROI"
        assert by_roi.cell(3, roi_headers["来源图片"]).value == "'=样品 B"
        assert by_roi.cell(3, roi_headers["状态"]).value == "失败"

        failures = workbook["失败明细"]
        failure_headers = _header_indexes(failures)
        assert failures.max_row == 2
        assert failures.cell(2, failure_headers["错误类型"]).value == "ValueError"
        assert failures.cell(2, failure_headers["错误消息"]).value == "'=危险公式"

        for sheet in workbook.worksheets:
            assert sheet.freeze_panes == "A2"
            assert sheet.auto_filter.ref
            assert not sheet.sheet_view.showGridLines
            assert all(
                cell.data_type != "f"
                for row in sheet.iter_rows()
                for cell in row
            )
    finally:
        workbook.close()


def test_batch_export_rows_preserve_item_order_and_resolve_scope_names() -> None:
    rows = build_analysis_batch_export_rows(
        _batch_result_with_failure(),
        document_names={"doc_1": "样品 A", "doc_2": "样品 B"},
        roi_names={"roi_1": "边缘 ROI", "roi_2": "失败 ROI"},
    )

    assert tuple(row.item_id for row in rows) == (
        "doc_1",
        "doc_1::roi::roi_1",
        "doc_2::roi::roi_2",
    )
    assert rows[0].scope == "整张图片"
    assert rows[1].scope == "ROI"
    assert rows[1].document_name == "样品 A"
    assert rows[1].roi_name == "边缘 ROI"
    assert rows[1].scalar_report
    assert rows[2].error_message == "=危险公式"


def test_multi_tool_batch_expands_each_execution_with_auditable_step(
    tmp_path: Path,
) -> None:
    result = _multi_tool_batch_result()
    rows = build_analysis_batch_export_rows(
        result,
        document_names={"doc_multi": "多工具样品"},
    )

    assert len(result.item_results) == 1
    assert len(rows) == 2
    assert tuple(row.item_id for row in rows) == ("doc_multi", "doc_multi")
    assert tuple(row.step_label for row in rows) == ("1/2", "2/2")
    assert tuple(row.tool_name for row in rows) == (
        "纤维方向性",
        "Haralick GLCM 纹理",
    )

    exported = export_analysis_batch_workbook(
        result,
        tmp_path / "multi.xlsx",
        document_names={"doc_multi": "多工具样品"},
    )
    assert exported.success
    assert exported.item_count == 1
    assert exported.success_count == 1
    workbook = load_workbook(exported.path, data_only=False)
    try:
        sheet = workbook["逐图片"]
        headers = _header_indexes(sheet)
        assert sheet.max_row == 3
        assert tuple(
            sheet.cell(row, headers["配方步骤"]).value
            for row in (2, 3)
        ) == ("1/2", "2/2")
        assert tuple(
            sheet.cell(row, headers["分析工具"]).value
            for row in (2, 3)
        ) == ("纤维方向性", "Haralick GLCM 纹理")
        assert workbook["总览"]["E2"].value == 1
        assert workbook["总览"]["J2"].value == 2
    finally:
        workbook.close()


def test_empty_cancelled_batch_still_exports_complete_empty_workbook(
    tmp_path: Path,
) -> None:
    result = AnalysisBatchResult(
        request_id="batch-cancelled",
        generation=4,
        recipe_id="directionality-v2",
        item_results=(),
        cancelled=True,
    )

    exported = AnalysisExportService().export_batch_workbook(
        result,
        tmp_path / "cancelled.xlsx",
    )

    assert exported.success
    workbook = load_workbook(exported.path, data_only=False)
    try:
        assert workbook["总览"]["D2"].value == "已取消"
        assert workbook["总览"]["E2"].value == 0
        assert workbook["逐图片"].max_row == 1
        assert workbook["逐 ROI"].max_row == 1
        assert workbook["失败明细"].max_row == 1
    finally:
        workbook.close()


def test_batch_export_rejects_inconsistent_success_without_writing(
    tmp_path: Path,
) -> None:
    inconsistent = AnalysisBatchResult(
        request_id="batch-invalid",
        generation=1,
        recipe_id="directionality-v2",
        item_results=(
            AnalysisBatchItemResult(
                item_id="doc_1",
                display_name="图像一",
                success=True,
            ),
        ),
    )
    target = tmp_path / "invalid.xlsx"

    exported = AnalysisExportService().export_batch_workbook(
        inconsistent,
        target,
    )

    assert not exported.success
    assert exported.failure_code is AnalysisExportFailureCode.ENCODE_FAILED
    assert not target.exists()


def test_batch_export_atomic_failure_preserves_existing_target(
    tmp_path: Path,
    monkeypatch,
) -> None:
    target = tmp_path / "batch.xlsx"
    target.write_bytes(b"existing-workbook")

    def fail_write(path, payload) -> None:
        assert Path(path) == target
        assert payload
        raise OSError("injected batch write failure")

    monkeypatch.setattr(
        "fdm.services.analysis_export.atomic_write_bytes",
        fail_write,
    )

    exported = AnalysisExportService().export_batch_workbook(
        _batch_result_with_failure(),
        target,
    )

    assert not exported.success
    assert (
        exported.failure_code
        is AnalysisExportFailureCode.ATOMIC_COMMIT_FAILED
    )
    assert target.read_bytes() == b"existing-workbook"
